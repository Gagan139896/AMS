import datetime
import math
import re
import time
from telnetlib import EC

import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.common.exceptions import TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global ClickCounter

  TestName = "test_CreateBreakdownElements"
  description = "This test scenario is to verify all the Elements present at Create Breakdown page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_CreateBreakdownElements"
  global Exe
  Exe="Yes"
  Directory = 'test_CreateBreakdown/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------
      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 1
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = ("D:/AMS/AmsTest/test_AmsActions/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active

        try:
            # ---------------------------Verify Create and Breakdown icon click-----------------------------
            PageName = "Create and BreakDown icon"
            Ptitle1 = "New    BreakDown   "
            try:
                driver.find_element_by_xpath("//li[@data-test-id='201812201359010458611']").click()
                time.sleep(2)
                driver.find_element_by_xpath("//li[@title='BreakDown']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                PageTitle1 = driver.find_element_by_xpath("//h2[text()='New    BreakDown   ']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present in left menu and able to click")
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            #--------------------------------------------------------------------------------------------------

            # ---------------------------Verify Create new breakdown window heading-----------------------------
            PageName = "Create new breakdown window heading"
            Ptitle1 = "New    BreakDown   "
            try:
                PageTitle1 = driver.find_element_by_xpath("//h2[text()='New    BreakDown   ']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present and text found is : " + PageTitle1)
                TestResult.append(PageName + " is present and text found is : " + PageTitle1)
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Asset Type text field label-----------------------------
            PageName = "Asset Type text field label"
            Ptitle1 = "Asset Type"
            try:
                PageTitle1 = driver.find_element_by_xpath("//label[@data-test-id='202203290036520952295-Label']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present and text found is : " + PageTitle1)
                TestResult.append(PageName + " is present and text found is : " + PageTitle1)
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Asset Name text field label-----------------------------
            PageName = "Asset Name text field label"
            Ptitle1 = "Asset Name"
            try:
                PageTitle1 = driver.find_element_by_xpath("//label[@data-test-id='202203290036520951686-Label']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present and text found is : " + PageTitle1)
                TestResult.append(PageName + " is present and text found is : " + PageTitle1)
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Breakdown Type DDL label-----------------------------
            PageName = "Breakdown Type dropdown label"
            Ptitle1 = "Breakdown Type"
            try:
                PageTitle1 = driver.find_element_by_xpath("//label[text()='Breakdown Type']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present and text found is : " + PageTitle1)
                TestResult.append(PageName + " is present and text found is : " + PageTitle1)
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Issue Description box label-----------------------------
            PageName = "Issue Description box label"
            Ptitle1 = "Issue Description"
            try:
                PageTitle1 = driver.find_element_by_xpath("//label[text()='Issue Description']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present and text found is : " + PageTitle1)
                TestResult.append(PageName + " is present and text found is : " + PageTitle1)
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Attach content button-----------------------------
            PageName = "Attach content button"
            Ptitle1 = "Attach content"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@id='pyFlowActionHTML']/span/div/div[2]//button").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present")
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Cancel button-----------------------------
            PageName = "Cancel button"
            Ptitle1 = "Cancel"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//button[@data-test-id='2014121801251706267665']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present")
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Submit button-----------------------------
            PageName = "Submit button"
            Ptitle1 = "Submit"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@data-test-id='202101201229000166203']/div[1]//button").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present")
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Close button-----------------------------
            PageName = "Close button"
            Ptitle1 = "Close"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//button[@data-test-id='201901261711080731178947']").get_attribute('title')
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present")
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ---------------------------Verify Minimize button-----------------------------
            PageName = "Minimize button"
            Ptitle1 = "Collapse"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//button[@data-test-id='201901261711080730177228']").get_attribute('title')
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present")
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present ")
                TestResultStatus.append("Fail")
            # --------------------------------------------------------------------------------------------------

            # ----------------------Verifying all the fields on Breakdown creation page-------------
            try:
                # ---------------Selecting asset type-----------------------------------------------
                PageName = "Asset type text field"
                try:
                    asset_type = driver.find_element_by_xpath("//input[@id='264249ac']")
                    asset_type.send_keys(sheet2.cell(3, 1).value)  # laptop
                    time.sleep(2)
                    asset_type.send_keys(Keys.DOWN)
                    asset_type.send_keys(Keys.ENTER)
                    time.sleep(3)
                    print("asset type done")
                    TestResult.append(PageName+ " is present and able to enter inputs")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName+ " is not present or not able to enter inputs")
                    TestResultStatus.append("Fail")
                # ---------------------------------------------------------------------------------------------

                # ---------------Selecting asset name-----------------------------------------------
                PageName = "Asset name text field"
                try:
                    asset_name = driver.find_element_by_xpath("//input[@id='50156d96']")
                    asset_name.send_keys(sheet2.cell(3, 2).value)
                    time.sleep(2)
                    asset_name.send_keys(Keys.DOWN)
                    time.sleep(2)
                    asset_name.send_keys(Keys.ENTER)
                    time.sleep(3)
                    TestResult.append(PageName+ " is present and able to enter inputs")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName+ " is not present or not able to enter inputs")
                    TestResultStatus.append("Fail")

                # ---------------Selecting hardcoded breakdown type--------------------------------------------------------
                PageName = "Breakdown type dropdown"
                try:
                    driver.find_element_by_xpath(
                        "//div[@class='content-item content-field item-3 remove-left-spacing remove-right-spacing flex required']//select[@id='9fa2b67b']").click()
                    BDown_type = driver.find_element_by_xpath(
                        "//div[@class='content-item content-field item-3 remove-left-spacing remove-right-spacing flex required']//select[@id='9fa2b67b']")
                    bd = Select(BDown_type)
                    bd.select_by_index(1)
                    time.sleep(5)
                    print("breakdown done")
                    TestResult.append(PageName+ " is present and able to select values from dropdown list")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName+ " is not present or not able to select values from dropdown list")
                    TestResultStatus.append("Fail")

                # ---------------Entering Issue Description-----------------------------------------------
                PageName = "Issue description text box"
                try:
                    driver.find_element_by_xpath(
                        "//div[@class='content-item content-field item-4 remove-left-spacing remove-right-spacing flex']//textarea[@id='9638b72b']").send_keys(
                        "Test description")
                    time.sleep(5)
                    TestResult.append(PageName+ " is present and able to enter inputs")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append(PageName+ " is not present or not able to enter inputs")
                    TestResultStatus.append("Fail")
            except Exception:
                pass

        except Exception:
            print("Presence of all elements on breakdown create page is not verified")
            TestResult.append("Presence of all elements on breakdown create page is not verified")
            TestResultStatus.append("Fail")
    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


