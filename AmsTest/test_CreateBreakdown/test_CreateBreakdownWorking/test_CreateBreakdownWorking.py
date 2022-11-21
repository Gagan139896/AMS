import datetime
import math
import re
import string
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_CreateBreakdownWorking"
  description = "This test scenario is to verify working of breakdown creation process"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_CreateBreakdownWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_CreateBreakdown/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = ("D:/AMS/AmsTest/test_AmsActions/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active

        try:
            #-----------------------Clicking on plus icon and create breakdown icon-------------
            Expected_Res = "BreakDown"
            try:
                driver.find_element_by_xpath("//li[@data-test-id='201812201359010458611']").click()
                time.sleep(2)
                driver.find_element_by_xpath("//li[@title='BreakDown']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break

                # ---------------Selecting asset type-----------------------------------------------
                asset_type = driver.find_element_by_xpath("//input[@id='264249ac']")
                asset_type.send_keys(sheet2.cell(3,1).value)  # laptop
                time.sleep(2)
                asset_type.send_keys(Keys.DOWN)
                asset_type.send_keys(Keys.ENTER)
                time.sleep(5)
                print("asset type done")
                #---------------------------------------------------------------------------------------------

                # ---------------Selecting asset name-----------------------------------------------
                asset_name = driver.find_element_by_xpath("//input[@id='50156d96']")
                asset_name.send_keys(sheet2.cell(3,2).value)
                time.sleep(2)
                asset_name.send_keys(Keys.DOWN)
                time.sleep(2)
                asset_name.send_keys(Keys.ENTER)
                time.sleep(5)

                # # ---------------Selecting random bd type-----------------------------------------------
                # bd_type = driver.find_element(By.XPATH,"//div[@class='content-item content-field item-3 remove-left-spacing remove-right-spacing flex required']//select[@id='9fa2b67b']")
                # time.sleep(2)
                # bd_type.click()
                # time.sleep(5)
                #
                # bd_types = driver.find_elements(By.XPATH,"//div[@class='content-item content-field item-3 remove-left-spacing remove-right-spacing flex required']//select[@id='9fa2b67b']/option")
                # len_bd_types = len(bd_types)
                # print(len_bd_types)
                #
                # random_bdtype = random.randint(1, len_bd_types)
                # print("random bd type" + str(random_bdtype))
                #
                # bd = Select(bd_type)
                # bd.select_by_index(random_bdtype)
                # time.sleep(5)
                # print("breakdown done")

                # ---------------Selecting hardcoded breakdown type--------------------------------------------------------
                driver.find_element_by_xpath("//div[@class='content-item content-field item-3 remove-left-spacing remove-right-spacing flex required']//select[@id='9fa2b67b']").click()
                BDown_type = driver.find_element_by_xpath("//div[@class='content-item content-field item-3 remove-left-spacing remove-right-spacing flex required']//select[@id='9fa2b67b']")
                bd = Select(BDown_type)
                bd.select_by_index(1)
                time.sleep(5)
                print("breakdown done")

                # ---------------Entering Issue Description-----------------------------------------------
                driver.find_element_by_xpath("//div[@class='content-item content-field item-4 remove-left-spacing remove-right-spacing flex']//textarea[@id='9638b72b']").send_keys(
                    "Test description")
                time.sleep(5)

                # --------------Clicking Attachment Button-------------------------------------------------

                attachment = driver.find_element_by_xpath("//div[@class='content-item content-field item-5 remove-left-spacing remove-right-spacing flex flex-row dataValueWrite']//button[@type='button'][normalize-space()='Attach content']").click()
                time.sleep(TimeSpeed)
                SelectFileBtn = driver.find_element_by_xpath("//input[@id='$PpyAttachmentPage$ppxAttachName']")
                SelectFileBtn.send_keys("C://Users/crochet-08/Downloads/Gagandeep-200-BIG.png")
                time.sleep(TimeSpeed)
                driver.find_element_by_id("ModalButtonSubmit").click()
                time.sleep(TimeSpeed)

                # ---------------Clicking Submit Button with valid details-------------------------------------------------
                driver.find_element_by_xpath("//button[normalize-space()='Submit']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                try:
                    Bd_Id = driver.find_element_by_xpath("//span[@data-test-id='20190510022618055338234']").text
                    print(Bd_Id)
                    sheet2.cell(11,1).value = Bd_Id
                    wb2.save(loc2)
                except Exception:
                    print()
                time.sleep(5)

                # # -------------To verify sucessful breakdown case creation by getting "Breakdwn" text------------
                #
                # Act_Res = driver.find_element_by_xpath("(//h1[normalize-space()='BreakDown'])[1]").text
                # print(Act_Res)
                # if Act_Res == Expected_Res:
                #     # -------------Storing test result--------------------------------------------------
                #     print("Breakdown case created successfully")
                #     TestResult.append("Breakdown case created successfully with valid details")
                #     TestResultStatus.append("Pass")
            except Exception:
                TestResult.append("Breakdown case creation process failed")
                TestResultStatus.append("Fail")
            print()
            time.sleep(5)

            # # ----------------Getting validation messages for mandatory fields on form submission--------------
            # Test_Component = "Submit Button"
            # Expected_Validation = " 'Value cannot be blank' "
            # try:
            #     # -----------Navigation to breakdown creation page---------------------------
            #     # ---------------to click on plus icon (create) and breakdown button-----------
            #     driver.find_element_by_xpath("//li[@data-test-id='201812201359010458611']").click()
            #     time.sleep(3)
            #     driver.find_element_by_xpath("//li[@data-test-id='201812201359010337524']").click()
            #     time.sleep(3)
            #
            #     # ----------Submittimg blank form------------------------------------
            #     driver.find_element_by_xpath("//button[normalize-space()='Submit']").click()
            #     print(Test_Component + " clicked")
            #     time.sleep(2)
            #
            #     # ----------Handling alert msg to correct flagged fields-----------------------------------------------------------------------------
            #     alert = Alert(driver)
            #     alert.accept()
            #     time.sleep(3)
            #
            #     # -----------Getting validation messages-------------------------------------------------------------------
            #     # ---------------asset type field------------------------------------------------------------------------
            #     asset_type_val = driver.find_element_by_xpath("//div[@id='$PpyWorkPage$pInitiatePM$pAssetTypeError']//span[@class='iconError dynamic-icon-error'][normalize-space()='Value cannot be blank']").text
            #     if asset_type_val in Expected_Validation:
            #         TestResult.append(
            #             Expected_Validation + " validation message is present when user clicked " + Test_Component + " without entering asset type")
            #         TestResultStatus.append("Pass")
            #         print(Expected_Validation + " validation message is present for Asset type field")
            #     else:
            #         print(Expected_Validation + " validation message is not present for Asset type field")
            #
            #     # ---------------asset name field--------------------------------------------------------------------------
            #     asset_name_val = driver.find_element_by_xpath("//div[@id='$PpyWorkPage$pInitiatePM$pNameError']//span[@class='iconError dynamic-icon-error'][normalize-space()='Value cannot be blank']").text
            #     if asset_name_val in Expected_Validation:
            #         TestResult.append(
            #             Expected_Validation + " validation message is present when user clicked " + Test_Component + " without entering asset name")
            #         TestResultStatus.append("Pass")
            #         print(Expected_Validation + " validation message is present for Asset name field")
            #     else:
            #         print(Expected_Validation + " validation message is not present for Asset name field")
            #
            #     # ---------------Breakdown type dropdown-----------------------------------------------
            #     breakdown_type_val = driver.find_element_by_xpath("//div[@id='$PpyWorkPage$pInitiatePM$pNameError']//span[@class='iconError dynamic-icon-error'][normalize-space()='Value cannot be blank']").text
            #     if breakdown_type_val in Expected_Validation:
            #         TestResult.append(
            #             Expected_Validation + " is present when user clicked " + Test_Component + " without entering Breakdown type")
            #         TestResultStatus.append("Pass")
            #         print(
            #             Expected_Validation + " validation message is present when user clicked for Breakdown type field")
            #     else:
            #         print(Expected_Validation + " validation message is not present for Breakdown type field")
            # except Exception:
            #     print("@@@---111---@@@")
            #     TestResult.append(
            #         Expected_Validation + " validation message is not present when user clicked " + Test_Component + " without entering Asset name, type and breakdown type")
            #     TestResultStatus.append("Fail")
            print()
            time.sleep(3)

            # # ----------------Scenario 3--------------------------------------------
            # # -----------To verify cancel button functionality to cancel breakdown case creation process ----------------------------
            # Test_Component = "Cancel Button"
            # try:
            #     # -----------Navigation to breakdown creation page---------------------------
            #     # ---------------to click on plus icon (create) and breakdown button-----------
            #     driver.refresh()
            #     time.sleep(5)
            #     driver.find_element_by_xpath("//li[@data-test-id='201812201359010458611']").click()
            #     time.sleep(2)
            #     driver.find_element_by_xpath("//li[@data-test-id='201812201359010337524']").click()
            #     for load in range(LONG_TIMEOUT):
            #         try:
            #             if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
            #                 time.sleep(0.5)
            #         except Exception:
            #             break
            #     # ----------Clicking cancel button-----------------------------------------------------------------------
            #     driver.find_element_by_xpath("//button[@title='Cancel']").click()
            #     time.sleep(2)
            #     # ----------Clicking delete button-----------------------------------------------------------------------
            #     driver.find_element_by_xpath("//button[@name='pyCloseCase_pyWorkPage_7']").click()
            #     time.sleep(2)
            #     TestResult.append(Test_Component + " is clickable")
            #     TestResultStatus.append("Pass")
            #     print(Test_Component + " is clickable")
            # except Exception:
            #     TestResult.append(Test_Component + " is not clickable")
            #     TestResultStatus.append("Fail")
            # print()
            # time.sleep(TimeSpeed)
            #
            # # ---------------------------------------------------------------------------------


        except Exception as err:
                print(err)
                TestResult.append("Breakdown creation process is not working correctly. Below error found\n"+str(err))
                TestResultStatus.append("Fail")
                pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


