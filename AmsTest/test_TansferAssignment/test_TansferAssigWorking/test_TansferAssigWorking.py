import datetime
import math
import random
import re
import string
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_TansferAssigWorking"
  description = "This test scenario is to verify working of Transfer Assignment page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_TansferAssigWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_TansferAssignment/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      #time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"

        try:
            # --------------------------------Transfer assignment icon click----------------------------------------------
            PageName = "Transfer assignment icon"
            Ptitle1 = "Transfer Assignment"
            try:
                driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[7]/a").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                driver.find_element_by_xpath("//input[@data-test-id='2015030515570700545405']").click()
                PageTitle1 = driver.find_element_by_xpath("//h2[text()='Transfer Assignment']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present in left menu and able to click")
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                print(PageName + " is not clickable")
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            # ----------------------------------------------------------------------------------------------------------

            # -------------Selecting Operator from DDL--------------------------------------------
            try:
                Operator_DDL_Count = driver.find_elements_by_xpath("//div[@data-test-id='202207141255050652255']//select/option")
                Operator_Rand = random.randrange(1, len(Operator_DDL_Count))
                print("Operator_Rand "+str(Operator_Rand))
                Operator_DDL = Select(driver.find_element_by_xpath("//div[@data-test-id='202207141255050652255']//select"))
                Operator_DDL.select_by_index(Operator_Rand)
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                TestResult.append("Operator is selected from opertor DDL")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append("Operator is not selected from opertor DDL")
                TestResultStatus.append("Fail")
                #---------------------------------------------------------------------------------

            # --------------------------------Transfer assignment process----------------------------------------------
            PageName = "Transfer assignment process"
            try:
                #--------Checking whether Operator worklist is expanded or not----------------------
                Operator_Worklist = driver.find_element_by_xpath("//div[@class='dl-accordion-btn']")
                Operator_Worklist1 = Operator_Worklist.get_attribute('aria-expanded')
                if Operator_Worklist1 == "true":
                    print("Operator_Worklist is : " +str(Operator_Worklist1))
                elif Operator_Worklist1 == "false":
                    print("Operator_Worklist is : " + str(Operator_Worklist1))
                    Operator_Worklist.click()
                try:
                    #-----------Clicking on select checkbox in grid-------------
                    Select_Checkbox = driver.find_element_by_xpath("//div[@class='oflowDivM ']/input[2]")
                    Select_Checkbox1 = Select_Checkbox.is_enabled()
                    print("Select_Checkbox1 " +str(Select_Checkbox1))
                    if Select_Checkbox1 == True:
                        print("11111")
                        time.sleep(3)
                        Select_Checkbox.click()
                        time.sleep(2)

                        #------------Getting "Transfer To" radio buttons count-----------------------
                        Transfer_To_Count = driver.find_elements_by_xpath("//div[@data-test-id='202207141356190159493']/div/span/input")
                        Transfer_To_Rand = random.randrange(1, len(Transfer_To_Count)+1)
                        print(Transfer_To_Rand)
                        driver.find_element_by_xpath("//div[@data-test-id='202207141356190159493']/div["+str(Transfer_To_Rand)+"]/span/input").click()

                        #--------------Selecting value from User DDL-----------
                        if Transfer_To_Rand == 1:
                            print("Transfer_To_Rand "+str(Transfer_To_Rand))
                            User_DDL_Count = driver.find_elements_by_xpath("//div[@data-test-id='202207130209200655486']/div[2]//select/option")
                            User_DDL_Rand = random.randrange(1, len(User_DDL_Count)+1)
                            print("User_DDL_Rand "+str(User_DDL_Rand))
                            User_DDL = Select(driver.find_element_by_xpath("//div[@data-test-id='202207130209200655486']/div[2]//select"))
                            User_DDL.select_by_index(User_DDL_Rand)
                            print("User is selected from user DDL")
                            TestResult.append("User is selected from user DDL")
                            TestResultStatus.append("Pass")

                        #-----------------Selecting value from Work Basket DDL-------------
                        else:
                            print("Transfer_To_Rand " + str(Transfer_To_Rand))
                            Work_Basket_Count = driver.find_elements_by_xpath("//select[@data-test-id='20220826151510017318']/option")
                            Work_Basket_Rand = random.randrange(1, len(Work_Basket_Count) + 1)
                            print("Work_Basket_Rand "+str(Work_Basket_Rand))
                            Work_Basket_DDL = Select(driver.find_element_by_xpath("//select[@data-test-id='20220826151510017318']"))
                            Work_Basket_DDL.select_by_index(Work_Basket_Rand)
                            print("Work basket is selected from work basket DDL")
                            TestResult.append("Work basket is selected from work basket DDL")
                            TestResultStatus.append("Pass")
                        driver.find_element_by_xpath("//button[@data-test-id='202207130209200657887']").click()
                        TestResult.append("Transfer button is clicked successfully")
                        TestResultStatus.append("Pass")
                        driver.refresh()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break
                        Progess_Msg = driver.find_element_by_xpath("//div[@data-test-id='202207181436320243932']")
                        Progess_Msg1 = Progess_Msg.is_displayed()
                        Progess_Msg2 = Progess_Msg.text
                        if Progess_Msg1 == True:
                            print("After clicking on transfer button. Below message is displayed : \n"+Progess_Msg2)
                            TestResult.append("After clicking on transfer button. Below message is displayed : \n"+Progess_Msg2)
                            TestResultStatus.append("Pass")
                except Exception:
                    try:
                        No_Work_Asgn = driver.find_element_by_xpath("//tr[@id='Grid_NoResults']//span").text
                        print("No_Work_Asgn "+str(No_Work_Asgn))
                        if No_Work_Asgn == "No work assigned":
                            TestResult.append("'" + No_Work_Asgn + "' : " "text found when there is no work assignment present in grid for selected operator")
                            TestResultStatus.append("Pass")
                    except Exception:
                        print("'No_Work_Asgn' exception")
                        TestResult.append("Transfer assignment is not working fine")
                        TestResultStatus.append("Fail")
            except Exception:
                print(PageName + " is not clickable")
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            # ----------------------------------------------------------------------------------------------------------

            # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            TestResult.append("Transfer assignment page is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


