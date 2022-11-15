import datetime
import math
import re
import string
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains, chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
import random
import os

from self import self


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_OrganizationChartWorking"
  description = "This test scenario is to verify working of Organization Chart page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_OrganizationChartWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_OrganizationChart/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = ("D:/AMS/AmsTest/test_AmsActions/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active

        try:
            #---------------To verify Organisation Chart icon click-----------------
            PageName = "Organisation chart icon"
            Ptitle1 = "Organisation chart"
            try:
                driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[5]/a").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                driver.find_element_by_xpath("//input[@title='Enter text to search']").click()
                time.sleep(2)
                PageTitle1 = driver.find_element_by_xpath("//h2[@class='header-title']").text
                if PageTitle1 == "Organisation Chart":
                    print(PageName + " is present in left menu and able to click")
                    TestResult.append(PageName + " is present in left menu and able to click")
                    TestResultStatus.append("Pass")
            except Exception:
                print(PageName + " is not clickable")
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            # -----------------------------------------------------------------------------------------

            #---------------Adding New Operator------------------------
            PageName = "Add New Operator button"
            Ptitle1 = "Organisation Chart"
            try:
                driver.find_element_by_xpath("//a[text()='Add New Operator']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                PageTitle1 = driver.find_element_by_xpath("//span[text()='        Create Operator       ']").is_displayed()
                if PageTitle1 == True:
                    print(PageName + " is clickable")
                    TestResult.append(PageName + " is clickable")
                    TestResultStatus.append("Pass")
            except Exception as errr:
                print(errr)
                print(PageName + " is not clickable")
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            #---------------------------------------------------------------

            #-----------Operator adding process--------------------------------
            PageName = "Operator adding process"
            try:
                # ---------------Organization selection------------------------------------------------------------
                Org_DDL_Count = driver.find_elements_by_xpath("//select[@data-test-id='202205261231470602516']/option")
                Org_Rand = random.randrange(1, len(Org_DDL_Count))
                Org_DDL = Select(driver.find_element_by_xpath("//select[@data-test-id='202205261231470602516']"))
                Org_DDL.select_by_index(Org_Rand)
                time.sleep(1)
                #-----------------------------------------------------------------------------------
                # ------------------------------ENTERING USER ID----------------------------------------------
                User_ID = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
                User_ID = "User_" + User_ID
                driver.find_element_by_xpath("//input[@data-test-id='202205101241230880204']").send_keys(User_ID)
                time.sleep(1)
                #-----------------------------------------------------------------------------------------------
                # ------------------------------ENTERING PASSWORD----------------------------------------------
                Password = ''.join(random.choices(string.digits, k=3))
                Password = "Pwd_" + Password
                driver.find_element_by_xpath("//input[@type='password']").send_keys(Password)
                time.sleep(1)
                #----------------------------------------------------------------------------------------------
                # ------------------------------ENTERING FULL NAME----------------------------------------------
                Full_Name = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
                Full_Name = "FName_" + Full_Name
                driver.find_element_by_xpath("//input[@data-test-id='202205101241230881398']").send_keys(Full_Name)
                time.sleep(1)
                #-----------------------------------------------------------------------------------------------
                # ---------------SELECTING ROLE------------------------------------------------------------
                Role_DDL_Count = driver.find_elements_by_xpath("//select[@data-test-id='202205101241230881520']/option")
                Role_Rand = random.randrange(1, len(Role_DDL_Count))
                print("Role_Rand is : "+str(Role_Rand))
                Role_DDL_Select = Select(driver.find_element_by_xpath("//select[@data-test-id='202205101241230881520']"))
                Role_DDL_Select.select_by_index(Role_Rand)
                time.sleep(5)
                #------------------------------------------------------------------------------------------
                # ----------------------------CLICKING ON SUBMIT BUTTON---------------------------
                driver.find_element_by_xpath("//button[@title='Submit']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
               #----------------------NEED TO WORK-----------------------------------
                try:
                    try:
                        PageTitle1 = driver.find_element_by_xpath("//span[@id='modaldialog_hd_title']")
                        if PageTitle1.is_displayed() == True:
                            time.sleep(5)
                            TestResult.append(PageName + " is not working. Because add operator window is not closed after submission the form")
                            TestResultStatus.append("Fail")
                    except Exception:
                        try:
                            print("Checking for alert message")
                            alert = driver.switch_to_alert()
                            alert_text = alert.text
                            time.sleep(5)
                            print(alert_text)
                            alert.accept()
                            TestResult.append(PageName + " is not working. Below alert message is found\n" + alert_text)
                            TestResultStatus.append("Fail")
                            print()
                        except Exception:
                            print(PageName + " is working fine")
                            TestResult.append(PageName + " is working fine")
                            TestResultStatus.append("Pass")
                            sheet2.cell(7, 1).value = User_ID
                            wb2.save(loc2)
                except Exception:
                    print("Inside exception")
                    pass
            except Exception as errrr:
                print(errrr)
                print("Organisation chart page is not working")
                TestResult.append("Organisation chart page is not working")
                TestResultStatus.append("Fail")
        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path + 'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


