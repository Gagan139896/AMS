import string
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver.support.select import Select
from sys import platform
import random
import os
import os.path



@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)


@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_SkillManagWorking"
  description = "This test scenario is to verify the Working of Elements at Skill Management page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_SkillManagWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_SkillManagement/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      #time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = ("D:/AMS/AmsTest/test_AmsActions/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active

        try:
            # ---------------------------Verify Skill Management icon click-----------------------------
            PageName = "Skill Management icon"
            Ptitle1 = "Skill Management"
            try:
                driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[6]/a").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                PageTitle1 = driver.find_element_by_xpath("//h1[text()='Skill Management']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present in left menu and able to click")
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
                driver.find_element_by_xpath("//input[@data-test-id='2015030515570700545405']").click()
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify '+ Add Skill' button click -----------------------------
            PageName = "'+ Add Skill' button"
            Ptitle1 = "        Add Skill       "
            try:
                driver.find_element_by_xpath("//button[@data-test-id='20160721092326035219972']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                PageTitle1 = driver.find_element_by_xpath("//span[@id = 'modaldialog_hd_title']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is clickable")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ----------------Adding skill-----------------
            PageName = "Skill adding process"
            Ptitle1 = ""
            try:
                #------------Selecting operator from Operator DDL---------------
                try:
                    Opertor_DDL_Count = driver.find_elements_by_xpath("//select[@data-test-id='202205310612020059753']/option")
                    Operator_Rand = random.randrange(1, len(Opertor_DDL_Count))
                    print(Operator_Rand)
                    Operator_DDL = Select(driver.find_element_by_xpath("//select[@data-test-id='202205310612020059753']"))
                    Operator_DDL.select_by_index(Operator_Rand)
                    time.sleep(2)
                    TestResult.append("Operator is selected from opertor DDL")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Operator is not selected from opertor DDL")
                    TestResultStatus.append("Fail")
                #-----------Entering skill------------------------------
                for aa in range(5):
                    letters = string.ascii_lowercase
                    returna = ''.join(random.choice(letters) for i in range(5))
                    FName = returna
                print(FName)
                LName = "_Skill"
                Skill_Name = FName + LName
                print(Skill_Name)

                try:
                    driver.find_element_by_xpath("//Label[@data-test-id='202205310612020060475-Label']").send_keys(Skill_Name)
                    time.sleep(2)
                    TestResult.append("Skill is entered successfully")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Not able to enter input in skill field")
                    TestResultStatus.append("Fail")
                #-------------Selecting Rating from DDL------------------------
                try:
                    Rating_DDL_Count = driver.find_elements_by_xpath("//select[@data-test-id='202208311501430532636']/option")
                    Rating_Rand = random.randrange(1, len(Rating_DDL_Count))
                    print(Rating_Rand)
                    Rating_DDL = Select(driver.find_element_by_xpath("//select[@data-test-id='202208311501430532636']"))
                    Rating_DDL.select_by_index(Rating_Rand)
                    time.sleep(2)
                    TestResult.append("Rating is selected from opertor DDL")
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("Rating is not selected from opertor DDL")
                    TestResultStatus.append("Fail")
                driver.find_element_by_xpath("// button[ @ id = 'ModalButtonSubmit']").click()
                TestResult.append("Submit button is clicked successfully")
                TestResultStatus.append("Pass")
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                try:
                    Error_Message = driver.find_element_by_xpath("//span[@id='ERRORMESSAGES_ALL']").is_displayed()
                    Error_Text = driver.find_element_by_xpath("//span[@id='ERRORMESSAGES_ALL']/ul/li[1]").text
                    if Error_Message == True:
                        driver.find_element_by_xpath("// button[ @ id = 'ModalButtonCancel']").click()
                        TestResult.append("Not able to add skill due to below  error \n"+str(Error_Text))
                        TestResultStatus.append("Fail")
                    else:
                        Page_Title = driver.find_element_by_xpath("//h1[@class='header-title']").is_displayed()
                        if Page_Title == True:
                            sheet2.cell(15, 1).value = Skill_Name
                            TestResult.append(PageName + "  is working fine")
                            TestResultStatus.append("Pass")
                            wb2.save(loc2)
                except Exception:
                    pass
            except Exception:
                TestResult.append(PageName + " is not working fine")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)

            # ---------------------------------------------------------------------------------

            # -----------------Verifying filters for skill table records------------------------
            try:
                #----Checking skill name in reffrence excel sheet------
                if sheet2.cell(15,1).value == None:
                    print("Skill name is not present is in reffrence excel sheet. Need to add skill first")
                    TestResult.append("Skill name is not present is in reffrence excel sheet. Need to add skill first")
                    TestResultStatus.append("Pass")
                else:
                    #-----Clicking on Skill filter------------
                    driver.find_element_by_xpath("//th[@data-test-id='202208021538000758244']//a").click()
                    #-----Code for loader--------
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    #-------Entering name in search field inside filter------------
                    driver.find_element_by_xpath("//input[@data-test-id='201411181100280377101613']").send_keys(sheet2.cell(15,1).value)
                    #------Clicking on apply button inside filter----------
                    driver.find_element_by_xpath("//ul[@class='pz-po-c-ul']/li[last()]/div/button[1]").click()
                    time.sleep(2)
                    #-------Getting text of filtered record----------
                    try:
                        Skill_Text = driver.find_element_by_xpath("//table[@pl_prop_class='Data-SkillManagement']/tbody/tr[2]/td[3]//span").text
                        if Skill_Text == sheet2.cell(15,1).value:
                            #----Clicking on delete button----------
                            driver.find_element_by_xpath("//button[@data-test-id='202208031200580364252']/i").click()
                            #------Clicking on submit button to verify record deletion----------
                            driver.find_element_by_xpath("//button[@id='ModalButtonSubmit']").click()
                            # -----Code for loader-----------
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            # -----Clicking on Skill filter after deleting skill record------------
                            driver.find_element_by_xpath("//th[@data-test-id='202208021538000758244']//a").click()
                            # -----Code for loader-----------
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            #-------Entering name in search field inside filter after deleting record------------
                            driver.find_element_by_xpath("//input[@data-test-id='201411181100280377101613']").send_keys(sheet2.cell(15, 1).value)
                            # ------Clicking on apply button inside filter after deleting record----------
                            driver.find_element_by_xpath("//ul[@class='pz-po-c-ul']/li[last()]/div/button[1]").click()
                            time.sleep(2)
                            try:
                                Skill_Text = driver.find_element_by_xpath("//table[@pl_prop_class='Data-SkillManagement']/tbody/tr[2]/td[3]//span").text
                                if Skill_Text != sheet2.cell(15, 1).value:
                                    for load in range(LONG_TIMEOUT):
                                        try:
                                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                                time.sleep(0.5)
                                        except Exception:
                                            break
                                    TestResult.append("Skill record is deleted successfully")
                                    TestResultStatus.append("Pass")
                                elif Skill_Text == sheet2.cell(15, 1).value:
                                    TestResult.append("Skill record deletion process is not working")
                                    TestResultStatus.append("Pass")
                            except Exception:
                                No_Record = driver.find_element_by_xpath("//table[@pl_prop_class='Data-SkillManagement']/tbody/tr[2]/td[1]//span").text
                                TestResult.append("The result after deleting the record and again applying filter is : " + str(No_Record))
                                sheet2.cell(15,1).value = None
                                wb2.save(loc2)
                        elif Skill_Text != sheet2.cell(15,1).value :
                            TestResult.append("Skill is not found in records after applying filter")
                            TestResultStatus.append("Fail")
                    except Exception as rrr:
                        No_Record1 = driver.find_element_by_xpath("//table[@pl_prop_class='Data-SkillManagement']/tbody/tr[2]/td[1]//span").text
                        TestResult.append("The result after applying filter is : "+str(No_Record1))
                        TestResultStatus.append("Fail")
                        print(rrr)
            except Exception:
                print()

        #---------------------------------------------------------------------------------------
        except Exception as err:
            print(err)
            TestResult.append("Skill management page is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName



        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


