import datetime
import random
import string
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from sys import platform
import os
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.select import Select


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global Rand_Num1
  global Rand_Num2
  global Asset_Rand
  global sheet2

  TestName = "test_AmsActionsWorking"
  description = "This test scenario is to verify all the Working of Elements at AMS Actions page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_AmsActionsWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_AmsActions/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      #driver.get("https://pegaqa.crochetech.com/prweb")
      #------Dev Env URL---------
      #driver.get("https://pegatest.crochetech.com/prweb/app/ACM/iCv8L5DiiUPcqfDwa-YAEw*/!STANDARD?pzPostData=-1389118839")
      #------QA Env URL----------
      driver.get("https://pegaqa.crochetech.com/prweb")
      #driver.get("http://ns.crochetech.com:8783/prweb/app/ACM")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      #time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    global Asset_Type_Name, Rand_Num2
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = (path + "/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active

        try:
            # ---------------------------Verify AMS Actions icon click-----------------------------
            PageName = "AMS Actions icon"
            Ptitle1 = "Add/Update Asset Type"
            try:
                #-----Clicking on AMS actions icon--------
                #driver.find_element_by_xpath("//li[@data-test-id='202203261757220141972']").click()
                driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[2]/a").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                time.sleep(TimeSpeed)
                driver.find_element_by_xpath("//div[@class=' flex content layout-content-inline_grid_double  content-inline_grid_double ']/child::div[1]//input[@id='pyGridActivePage']").click()
                PageTitle1 = driver.find_element_by_xpath("//h2[text()='Add/Update Asset Type']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Filter search in Asset Type table-----------------------------
            PageName = "Asset type table Filter"
            try:
                if sheet2.cell(3,1).value==None:
                    TestResult.append("Asset type name is not found in excel sheet. Need to add new asset type first")
                    TestResultStatus.append("Pass")
                    # ---------------------------Adding New Asset Type--------------------------------
                    PageName = "New asset type "

                    for aa in range(5):
                        letters = string.ascii_lowercase
                        returna = ''.join(random.choice(letters) for i in range(5))
                        FName = returna
                    print(FName)
                    LName = "_AssetType"
                    Asset_Type_Name = FName + LName
                    print(Asset_Type_Name)

                    try:
                        #-----Clicking on add asset type button-----------
                        driver.find_element_by_xpath("//a[text()='+ Asset Type']").click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break
                        time.sleep(TimeSpeed)
                        TestResult.append("'+ Asset Type' button is clicked successfully")
                        TestResultStatus.append("Pass")
                        #------Entering assert type name in asset type field-----------
                        driver.find_element_by_xpath("//input[@data-test-id='202203300846480686894']").send_keys(
                            Asset_Type_Name)
                        TestResult.append("Asset type name is entered")
                        TestResultStatus.append("Pass")
                        #-------Clicking on submit button----------
                        driver.find_element_by_xpath("//button[text()='  Submit ']").click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break
                        TestResult.append("Submit button is clicked")
                        TestResultStatus.append("Pass")
                        TestResult.append(PageName + " added successfully. Asset Type name is : " + Asset_Type_Name)
                        TestResultStatus.append("Pass")
                        Verify_Asset_Type = driver.find_element_by_xpath("//h2[text()='Add/Update Asset Type']").is_displayed()
                        if Verify_Asset_Type == True:
                            sheet2.cell(3, 1).value = Asset_Type_Name
                            wb2.save(loc2)
                    except Exception:
                        TestResult.append(PageName + "add process not working properly")
                        TestResultStatus.append("Fail")
                    print()
                    time.sleep(TimeSpeed)
                    # ---------------------------------------------------------------------------------

                    #-----------------Deleting Asset Type by applying filter-----------------------------
                else:
                    #------------------Clicking on filter---------------------
                    Search_Filter = driver.find_element_by_xpath("//th[@data-attribute-name='Asset Type']/div/span/a")
                    Search_Filter.click()
                    TestResult.append("Clicked on filter present in Asset type table")
                    TestResultStatus.append("Pass")

                    #--------------Entering value in Search text field----------
                    Search_Text_Field = driver.find_element_by_xpath("//input[@data-test-id='201411181100280377101613']")
                    Search_Text_Field.send_keys(sheet2.cell(3, 1).value)
                    TestResult.append("Data is entered in filter")
                    TestResultStatus.append("Pass")

                    #------------Clicking on apply button in filter--------------
                    Filter_Apply_Btn = driver.find_element_by_xpath("//ul[@class='pz-po-c-ul']/li[3]/div/button[1]")
                    Filter_Apply_Btn.click()
                    for load in range(LONG_TIMEOUT):
                        try:
                            if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                time.sleep(0.5)
                        except Exception:
                            break
                    TestResult.append("Apply button is clicked to perform filter search")
                    TestResultStatus.append("Pass")
                    time.sleep(2)

                    #---------Getting text asset type to be deleted--------------
                    Deleted_Asset_Type = driver.find_element_by_xpath(
                        "//table[@pl_prop_class='AMS-Data-AssetTypes']/tbody/tr[2]/td[1]//span").text
                    print("Deleted_Asset_Type " +Deleted_Asset_Type)
                    TestResult.append("Record found after applying filter is : "+Deleted_Asset_Type)
                    TestResultStatus.append("Pass")
                    time.sleep(2)
                    if Deleted_Asset_Type == "No items for the filters applied":
                        Search_Filter = driver.find_element_by_xpath("//th[@data-attribute-name='Asset Type']/div/span/a")
                        Search_Filter.click()
                        time.sleep(2)
                        driver.find_element_by_xpath("//a[@id='clearFilter']").click()
                        sheet2.cell(3,1).value=None
                        wb2.save(loc2)
                    else:
                        #-----------Clicking on delete icon-----------
                        driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-AssetTypes']/tbody/tr[2]/td[3]//i[@data-test-id='202203300855020903290']").click()

                        #--------Clicking on submit button to confirm delete------------
                        driver.find_element_by_xpath("//button[@id='ModalButtonSubmit']").click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break

                    #----------Verifying deleted record by applying filter again-----
                    try:
                        Search_Filter.click()
                        Search_Text_Field.send_keys(sheet2.cell(3, 1).value)
                        Filter_Apply_Btn.click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break
                        Deleted_Asset_Type2 = Deleted_Asset_Type.text
                        if Deleted_Asset_Type2 == "No items for the filters applied":
                            TestResult.append("Asset type deleted successfully from asset type table records")
                            TestResultStatus.append("Pass")
                            TestResult.append(PageName + " is working fine. Asset type found after applying filter is: " + Deleted_Asset_Type)
                            TestResultStatus.append("Pass")
                            Search_Filter.click()
                            driver.find_element_by_xpath("//a[@id='clearFilter']").click()
                    except Exception:
                        pass

                    #----Saving Asset type records in excel-----
                    sheet2.cell(3, 1).value = None
                    wb2.save(loc2)
            except Exception:
                pass
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify pagination clicks in Asset Type Table-----------------------------
            PageName = "Asset type table pagination"

            #--------Getting number of rows each page------
            try:
                #-------Getting total number of pages in asset type table-----
                Page_Count1 = driver.find_element_by_xpath("//div[@class=' flex content layout-content-inline_grid_double  content-inline_grid_double ']/child::div[1]//label[@data-test-id='20141007100658002115508']").text
                NumOfClicks1 = int(Page_Count1)
                print(Page_Count1)
                print(NumOfClicks1)

                counter=0
                #----For loop to page clicks----
                for i in range(1,NumOfClicks1+1):
                    AssTypNumRows1 = driver.find_elements_by_xpath("//table[@pl_prop_class='AMS-Data-AssetTypes']/tbody/tr")
                    NumRows1 = int(len(AssTypNumRows1))

                    #---For loop to get no. of records per page----
                    for ii in range(NumRows1):
                        if ii >= 5:
                            pass
                        else:
                            try:
                                Asset_Text = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-AssetTypes']/tbody/tr[" + str(ii+2) + "]/td[1]//span").is_displayed()
                                time.sleep(0.5)
                                if Asset_Text == True:
                                    counter = counter + 1
                            except Exception:
                                break

                    print("No. of records on page no. " + str(i) + " are : " + str(counter))
                    driver.find_element_by_xpath("//div[@class=' flex content layout-content-inline_grid_double  content-inline_grid_double ']/child::div[1]//button[@title='Next Page']").click()
                print(counter)
                #print(PageName + "  is working fine for " + str(NumOfClicks1) + " pages and No. of records found are : " + str(NumOfAssetType))
                TestResult.append(PageName + "  is working fine for " + str(NumOfClicks1) + " pages and No. of records found are : " + str(counter))
                TestResultStatus.append("Pass")
                sheet2.cell(3,4).value = counter
                wb2.save(loc2)
            except Exception as err:
                print(err)
                TestResult.append(PageName + " is not working")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Filter search in Asset table-----------------------------
            PageName = "Asset table Filter"

            #----Checking asset type name in excel sheet---------
            if sheet2.cell(3,1)==None:
                TestResult.append("Asset type is not found in excel sheet. Need to skip Asset add process")
                TestResultStatus.append("Pass")
            else:
                try:
                    #---Checking asset name in excel sheet----
                    if sheet2.cell(3, 2).value == None:
                        TestResult.append("Asset name is not found in excel sheet. Need to add new asset first")
                        TestResultStatus.append("Pass")

                        # ---------------------------Adding New Asset--------------------------------
                        PageName1 = "New asset "
                        for aa in range(5):
                            letters = string.ascii_lowercase
                            returna = ''.join(random.choice(letters) for i in range(5))
                            FName = returna
                        print(FName)
                        LName = "_Asset"
                        Asset_Name = FName + LName
                        print(Asset_Name)

                        try:
                            #-----Clicking on +Asset icon------
                            driver.find_element_by_xpath("//a[text()='+ Asset']").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            time.sleep(TimeSpeed)
                            #----Clicking on submit button-------
                            driver.find_element_by_xpath("//button[@data-test-id='2014121801251706289770' and text()='Submit']").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            time.sleep(2)
                            Create_Asset = driver.find_elements_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div")
                            Create_Asset = len(Create_Asset)
                            #----For loop to fill the add asset form-----
                            for ca in range(Create_Asset):
                                #---------Selecting Organization---------
                                if ca==0:
                                    Org_DDL = Select(driver.find_element_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div["+str(ca+1)+"]//select"))
                                    Org_DDL.select_by_index(1)
                                    time.sleep(2)
                                #------Selecting asset type-------------
                                elif ca==3:
                                    AssetType = driver.find_element_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div["+str(ca+1)+"]//input[@type='text']")
                                    AssetType.send_keys(sheet2.cell(3,1).value)
                                    AssetType.send_keys(Keys.DOWN)
                                    AssetType.send_keys(Keys.ENTER)
                                    time.sleep(2)
                                #------Entering asset--------------
                                elif ca==4:
                                    Asset = driver.find_element_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div["+str(ca+1)+"]//input[@type='text']")
                                    Asset.send_keys(Asset_Name)
                                    time.sleep(2)
                                #---------Entering specifications----------
                                elif ca==5:
                                    Specifications = driver.find_element_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div["+str(ca+1)+"]//input[@type='text']")
                                    Specifications.send_keys(FName)
                                    time.sleep(2)
                                #-------Selceting "Whether in use"-----------
                                elif ca==6:
                                    WIU_DDL = Select(driver.find_element_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div["+str(ca+1)+"]//select"))
                                    WIU_DDL.select_by_index(1)
                                    time.sleep(2)
                                #-----Entering asset address--------
                                elif ca==8:
                                    Asset_Address = driver.find_element_by_xpath("//div[@data-test-id='20200506135148097754']//div[@class='layout-body clearfix  ']/div/div["+str(ca+1)+"]//input[@type='text']")
                                    Asset_Address.send_keys("xyz")
                                    time.sleep(2)

                            driver.find_element_by_xpath("//button[@data-test-id='2014121801251706289770' and text()='Submit']").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            TestResult.append(PageName1 + " added successfully. Asset name is : " + Asset_Name)
                            TestResultStatus.append("Pass")
                            time.sleep(3)
                            VerifyRecord = driver.find_element_by_xpath("//div[@data-test-id='201712290453170848504']//div[@data-test-id='201808160754420438797']/div[1]/span[text()='AMC and Breakdown Cases List']").is_displayed()
                            if VerifyRecord == True:
                                sheet2.cell(3, 2).value = Asset_Name
                                wb2.save(loc2)
                            driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[2]/a").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                        except Exception:
                            TestResult.append(PageName1+ "add process not working properly")
                            TestResultStatus.append("Fail")
                        print()
                        time.sleep(5)
                        # ---------------------------------------------------------------------------------

                    else:
                        Search_Filter = driver.find_element_by_xpath("//th[@data-attribute-name='Asset']/div/span/a")
                        Search_Filter.click()
                        Search_Text_Field = driver.find_element_by_xpath("//input[@data-test-id='201411181100280377101613']")
                        Search_Text_Field.send_keys(sheet2.cell(3, 2).value)
                        Filter_Apply_Btn = driver.find_element_by_xpath("//ul[@class='pz-po-c-ul']/li[3]/div/button[1]")
                        Filter_Apply_Btn.click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break

                        # ---------Getting text asset to be deleted--------------
                        Deleted_Asset = driver.find_element_by_xpath(
                            "//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr[2]/td[1]//span").text
                        TestResult.append("Record found after applying filter is : " + Deleted_Asset)
                        TestResultStatus.append("Pass")
                        time.sleep(2)

                        time.sleep(2)
                        if Deleted_Asset == "No items for the filters applied":
                            Search_Filter = driver.find_element_by_xpath("//th[@data-attribute-name='Asset']/div/span/a")
                            Search_Filter.click()
                            time.sleep(2)
                            driver.find_element_by_xpath("//a[@id='clearFilter']").click()
                            sheet2.cell(3, 2).value = None
                            wb2.save(loc2)
                        else:
                            # -----------Clicking on delete icon-----------
                            driver.find_element_by_xpath(
                                "//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr[2]/td[4]//i[@data-test-id='202203300135520890667']").click()
                            # --------Clicking on submit button to confirm delete------------
                            driver.find_element_by_xpath("//button[@id='ModalButtonSubmit']").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break

                        # ----------Verifying deleted asset record by applying filter again-----
                        try:
                            Search_Filter.click()
                            Search_Text_Field.send_keys(sheet2.cell(3, 1).value)
                            Filter_Apply_Btn.click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            Deleted_Asset2 = Deleted_Asset.text
                            if Deleted_Asset2 == "No items for the filters applied":
                                TestResult.append("Asset type deleted successfully from asset type table records")
                                TestResultStatus.append("Pass")
                                TestResult.append(PageName + " is working fine. Asset type found after applying filter is: " + Deleted_Asset)
                                TestResultStatus.append("Pass")
                                Search_Filter.click()
                                driver.find_element_by_xpath("//a[@id='clearFilter']").click()
                        except Exception:
                            pass
                        sheet2.cell(3, 2).value=None
                        wb2.save(loc2)
                except Exception:
                    pass
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify pagination clicks in Asset Table-----------------------------
            PageName = "Asset table Pagination"
            try:
                Page_Count2 = driver.find_element_by_xpath(
                    "//div[@class=' flex content layout-content-inline_grid_double  content-inline_grid_double ']/child::div[2]//label[@data-test-id='20141007100658002115508']").text
                NumOfClicks2 = int(Page_Count2)
                print(Page_Count2)
                print(NumOfClicks2)

                Is_Yes = 0
                Is_No = 0
                for i1 in range(1,NumOfClicks2+1):
                    AssNumRows2 = driver.find_elements_by_xpath("//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr")
                    NumRows2 = len(AssNumRows2)

                    for ii1 in range(NumRows2):
                        if ii1 >= 5:
                            pass
                        else:
                            try:
                                Asset_Status = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr[" + str(ii1 + 2) + "]/td[3]//span").text
                                time.sleep(0.5)
                                if Asset_Status == "Yes":
                                    Is_Yes = Is_Yes + 1
                                elif Asset_Status == "No":
                                    Is_No = Is_No + 1
                            except Exception:
                                break
                    print("No. of records on page no. " + str(i1) + " are : " + str(Is_Yes+Is_No))
                    driver.find_element_by_xpath("//div[@class=' flex content layout-content-inline_grid_double  content-inline_grid_double ']/child::div[2]//button[@title='Next Page'] ").click()
                TotalAssets = Is_Yes  + Is_No
                TestResult.append(PageName + "  is working fine for " +str(NumOfClicks2)+ " pages and No. of records found are : " +str(TotalAssets))
                TestResultStatus.append("Pass")
                sheet2.cell(3, 5).value = TotalAssets
                sheet2.cell(3, 7).value = Is_Yes
                sheet2.cell(3, 8).value = Is_No
                wb2.save(loc2)
            except Exception as errr:
                print(errr)
                TestResult.append(PageName + " is not working")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ------Scrolling Page----------------------------------------------------------
            for scrolldown in range(1, 10):
                time.sleep(2)
                try:
                    driver.find_element_by_xpath(
                        "//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']//td[7]//button")
                    break
                except Exception:
                    # ActionChains(driver).key_down(Keys.).perform()
                    print("Inside Excep")
                    ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                    print("Page Down")
                    pass
            # -----------------------------------------------------------------------------------------------

            # #---------Verifying validation messages for required fields during Activity Add process-----------
            # PageName = "Validation message"
            # Ptitle1 = ""
            #
            # try:
            #     driver.find_element_by_xpath("//a[text()='+ Activity']").click()
            #     driver.find_element_by_xpath("//button[text()='  Submit ']").click()
            #     alert = Alert(driver)
            #     alert.accept()
            #     for er in range(1,4):
            #         try:
            #             Field_Name = driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div["+str(er)+"]/label").text
            #             Val_Msg = driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div["+str(er)+"]//span[@class='iconError dynamic-icon-error']").text
            #             TestResult.append(PageName + " is present ("+Val_Msg+") for '"+Field_Name+"' field. When user click submit without entering any input")
            #             print(PageName + " is present ("+Val_Msg+") for '"+Field_Name+"' field. When user click submit without entering any input")
            #             TestResultStatus.append("Pass")
            #         except Exception as error:
            #             print(error)
            #             TestResult.append(PageName + " is not present for '"+Field_Name+"' field. When user click submit without entering any input")
            #             TestResultStatus.append("Fail")
            #     driver.find_element_by_xpath("//button[text()='  Cancel ']").click()
            # except Exception:
            #     TestResult.append(
            #         PageName + "are not present for required fields")
            #     TestResultStatus.append("Fail")
            #     print("")
            # time.sleep(3)
            # #--------------------------------------------------------------------------------------------------
            # # ------Scrolling Page----------------------------------------------------------
            # for scrolldown in range(1, 10):
            #     time.sleep(2)
            #     try:
            #         driver.find_element_by_xpath(
            #             "//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']//td[7]//button")
            #         break
            #     except Exception:
            #         # ActionChains(driver).key_down(Keys.).perform()
            #         print("Inside Excep")
            #         ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
            #         print("Page Down")
            #         pass
            # # -----------------------------------------------------------------------------------------------
            # #-------------------------Getting error messages------------------------------------------------
            # PageName = "Error message "
            #
            # try:
            #     driver.find_element_by_xpath("//a[text()='+ Activity']").click()
            #     time.sleep(2)
            #     Asset_Select = driver.find_element_by_xpath(
            #         "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[1]/div/input[@data-test-id='202203300846480686894']")
            #     Asset_Select.send_keys("Laptop")
            #     Asset_Select.send_keys(Keys.DOWN)
            #     Asset_Select.send_keys(Keys.ENTER)
            #     driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[2]//input").send_keys("Test_Activity")
            #     driver.find_element_by_xpath(
            #         "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[3]//input").send_keys(3)
            #     driver.find_element_by_xpath(
            #         "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[4]//input").send_keys(2)
            #     Skill_Select = Select(driver.find_element_by_xpath(
            #         "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[5]//select"))
            #     Skill_Select.select_by_index(2)
            #     try:
            #         Add_Item = driver.find_element_by_xpath("//a[@data-test-id='2015092514042309017947']")
            #         Add_Item.click()
            #         # Add_Item.click()
            #         Task_Name = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-PlanningTasks']//input[@class='leftJustifyStyle']")
            #         Task_Name.clear()
            #         Task_Name.send_keys("Test Task")
            #         Due_In_Days = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-PlanningTasks']//input[@class='rightJustifyStyle']")
            #         Due_In_Days.clear()
            #         Due_In_Days.send_keys(3)
            #         Action_DDL = Select(driver.find_element_by_xpath(
            #             "//select[@data-test-id='2016072109335505834280']"))
            #         Action_DDL.select_by_index(2)
            #     except Exception:
            #         pass
            #     driver.find_element_by_xpath("//button[text()='  Submit ']").click()
            #     Err_Message = driver.find_element_by_xpath("//span[@id='ERRORMESSAGES_ALL']/ul/li").text
            #     TestResult.append(PageName + "found is : "+ Err_Message +". When 'Due in days value is greater than 'Planning in days' value")
            #     TestResultStatus.append("Pass")
            #     driver.find_element_by_xpath("//button[text()='  Cancel ']").click()
            # except Exception as eccc:
            #     print(eccc)
            #     TestResult.append(" No "+ PageName + " is found when 'Due in days' value is greater than 'Planning in days' value")
            #     TestResultStatus.append("Fail")
            #     print("@@@@@")
            # print()
            # time.sleep(3)
            #
            # #-----------------------------------------------------------------------------------------------
            #
            # # ------Scrolling Page----------------------------------------------------------
            # for scrolldown in range(1, 10):
            #     time.sleep(1)
            #     try:
            #         driver.find_element_by_xpath(
            #             "//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']//td[7]//button")
            #         break
            #     except Exception:
            #         # ActionChains(driver).key_down(Keys.).perform()
            #         print("Inside Excep")
            #         ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
            #         print("Page Down")
            #         pass
            # # -----------------------------------------------------------------------------------------------
            #
            # # ---------------------------Adding New Activities--------------------------------
            # PageName = "New activity "
            # Ptitle1 = ""
            # list1 = [1, 2, 3, 4, 5, 6, 7, 8]
            #
            # for aa in range(5):
            #     letters = string.ascii_lowercase
            #     returna = ''.join(random.choice(letters) for i in range(5))
            #     FName = returna
            # print(FName)
            # LName = "_Activity"
            # Activity_Name = FName + LName
            # print(Activity_Name)
            #
            # try:
            #     driver.find_element_by_xpath("//a[text()='+ Activity']").click()
            #     time.sleep(TimeSpeed)
            #
            #     Asset_Select = driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[1]//input")
            #     Asset_Select.send_keys(Asset_Type_Name)
            #     time.sleep(2)
            #     Asset_Select.send_keys(Keys.DOWN)
            #     Asset_Select.send_keys(Keys.ENTER)
            #     time.sleep(2)
            #     driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[2]//input").send_keys(Activity_Name)
            #
            #     for i in range(1,100):
            #         Rand_Num1 = random.choice(list1)
            #         Rand_Num2 = random.choice(list1)
            #         if Rand_Num1>Rand_Num2:
            #             driver.find_element_by_xpath(
            #                 "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[3]//input").send_keys(
            #                 Rand_Num1)
            #             driver.find_element_by_xpath(
            #                 "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[4]//input").send_keys(
            #                 Rand_Num2)
            #             break
            #     try:
            #         Skill_Count = driver.find_elements_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[5]//select/option")
            #         Skill_Rand = random.randrange(1,len(Skill_Count))
            #         print(Skill_Rand)
            #
            #         Skill_Select = Select(driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[5]//select"))
            #         Skill_Select.select_by_index(Skill_Rand)
            #         Add_Item = driver.find_element_by_xpath("//a[@data-test-id='2015092514042309017947']")
            #         Add_Item.click()
            #         Task_Name = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-PlanningTasks']//input[@class='leftJustifyStyle']")
            #         Task_Name.clear()
            #         Task_Name.send_keys(FName)
            #         Due_In_Days = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-PlanningTasks']//input[@class='rightJustifyStyle']")
            #         Due_In_Days.clear()
            #         Due_In_Days.send_keys(Rand_Num2-1)
            #         Action_DDL = Select(driver.find_element_by_xpath(
            #             "//select[@data-test-id='2016072109335505834280']"))
            #         Action_DDL.select_by_index(2)
            #         driver.find_element_by_xpath("//button[text()='  Submit ']").click()
            #         TestResult.append(PageName + "added successfully. Activity name is : " + Activity_Name)
            #         TestResultStatus.append("Pass")
            #         sheet2.cell(3, 3).value = Activity_Name
            #         wb2.save(loc2)
            #     except Exception:
            #         pass
            # except Exception:
            #     TestResult.append(PageName + "add process not working properly")
            #     TestResultStatus.append("Fail")
            # print()
            # time.sleep(5)
            # # ---------------------------------------------------------------------------------

            # # ---------------------------Verify pagination clicks in Activity Table-----------------------------
            # PageName = "Activities table Pagination"
            # ActNumRows = driver.find_elements_by_xpath("//table[@pl_prop_class='AMS-Data-Activities']/tbody/tr")
            # NumRows3 = len(ActNumRows)
            # try:
            #     Page_Count3 = driver.find_element_by_xpath(
            #         "//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']/tbody/tr/td[6]").text
            #     NumOfClicks3 = int(Page_Count3)
            #     print(Page_Count3)
            #     print(NumOfClicks3)
            #     for i in range(1,NumOfClicks3):
            #         for ii in range(1, NumRows3):
            #             Asset_Name_Text = driver.find_element_by_xpath(
            #                 "//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr[" + str(ii + 1) + "]/td[2]//a").text
            #             time.sleep(0.5)
            #             print(Asset_Name_Text)
            #             if Asset_Name_Text == "Asset_Name":
            #                 # driver.find_element_by_xpath(
            #                 #     "//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr[" + str(ii + 1) + "]/td[5]//i[@data-test-id='202203300135520890667']").click()
            #                 # time.sleep(2)
            #                 #driver.find_element_by_xpath("//button[@id='ModalButtonSubmit']").click()
            #                 Asset_find = "True"
            #                 break
            #             print("Bool1 is " + Asset_find)
            #             if Asset_find == "True":
            #                     break
            #             else:
            #              driver.find_element_by_xpath(
            #             "//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']//td[7]//button").click()
            #
            #     TestResult.append(PageName + "  is working fine")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " is not working")
            #     TestResultStatus.append("Fail")
            # print()
            # # ---------------------------------------------------------------------------------

            # ---------------------------Verify Filter search in Activity table-----------------------------
            PageName = "Activity table Filter"
            list1 = [1, 2, 3, 4, 5, 6, 7, 8]

            # ----Checking asset type name in excel sheet---------
            if sheet2.cell(2, 1) == None:
                TestResult.append("Asset type is not found in excel sheet. Need to skip Activity add process")
                TestResultStatus.append("Pass")
            else:
                try:
                    # ---Checking activity name in excel sheet----
                    if sheet2.cell(3, 3).value == None:
                        TestResult.append("Activity name is not found in excel sheet. Need to add new activity first")
                        TestResultStatus.append("Pass")

                        # ---------------------------Adding New Asset--------------------------------
                        PageName1 = "New activity "
                        for aa in range(5):
                            letters = string.ascii_lowercase
                            returna = ''.join(random.choice(letters) for i in range(5))
                            FName = returna
                        print(FName)
                        LName = "_Activity"
                        Activity_Name = FName + LName
                        print(Activity_Name)

                        try:
                            driver.find_element_by_xpath("//a[text()='+ Activity']").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            time.sleep(TimeSpeed)

                            Asset_Select = driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[1]//input[@data-test-id='202203300846480686894']")
                            Asset_Select.send_keys(Asset_Type_Name)
                            time.sleep(2)
                            Asset_Select.send_keys(Keys.DOWN)
                            Asset_Select.send_keys(Keys.ENTER)
                            time.sleep(2)
                            driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[2]//input").send_keys(Activity_Name)

                            for i in range(1,100):
                                Rand_Num1 = random.choice(list1)
                                Rand_Num2 = random.choice(list1)
                                if Rand_Num1>Rand_Num2:
                                    driver.find_element_by_xpath(
                                        "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[3]//input").send_keys(
                                        Rand_Num1)
                                    driver.find_element_by_xpath(
                                        "//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[4]//input").send_keys(
                                        Rand_Num2)
                                    break
                            try:
                                Skill_Count = driver.find_elements_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[5]//select/option")
                                Skill_Rand = random.randrange(1,len(Skill_Count))
                                print(Skill_Rand)

                                Skill_Select = Select(driver.find_element_by_xpath("//div[@id='modaldialog_con']//div[@data-test-id='201808120943310957798']/div[5]//select"))
                                Skill_Select.select_by_index(Skill_Rand)
                                Add_Item = driver.find_element_by_xpath("//a[@data-test-id='2015092514042309017947']")
                                Add_Item.click()
                                Task_Name = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-PlanningTasks']//input[@class='leftJustifyStyle']")
                                Task_Name.clear()
                                Task_Name.send_keys(FName)
                                Due_In_Days = driver.find_element_by_xpath("//table[@pl_prop_class='AMS-Data-PlanningTasks']//input[@class='rightJustifyStyle']")
                                Due_In_Days.clear()
                                Due_In_Days.send_keys(Rand_Num2-1)
                                Action_DDL = Select(driver.find_element_by_xpath("//select[@data-test-id='2016072109335505834280']"))
                                Action_DDL.select_by_index(2)
                                driver.find_element_by_xpath("//button[text()='  Submit ']").click()
                                for load in range(LONG_TIMEOUT):
                                    try:
                                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                            time.sleep(0.5)
                                    except Exception:
                                        break
                                TestResult.append(PageName1 + "added successfully. Activity name is : " + Activity_Name)
                                TestResultStatus.append("Pass")

                                sheet2.cell(3, 3).value = Activity_Name
                                wb2.save(loc2)
                            except Exception:
                                pass
                        except Exception:
                            TestResult.append(PageName + "add process not working properly")
                            TestResultStatus.append("Fail")
                        print()
                        time.sleep(5)
                        # ---------------------------------------------------------------------------------
                    else:
                        Search_Filter = driver.find_element_by_xpath(
                            "//th[@data-attribute-name='Activity']/div/span/a")
                        Search_Filter.click()
                        Search_Text_Field = driver.find_element_by_xpath(
                            "//input[@data-test-id='201411181100280377101613']")
                        Search_Text_Field.send_keys(sheet2.cell(3, 3).value)
                        Filter_Apply_Btn = driver.find_element_by_xpath(
                            "//ul[@class='pz-po-c-ul']/li[3]/div/button[1]")
                        Filter_Apply_Btn.click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break

                        # ---------Getting text activity to be deleted--------------
                        Deleted_Activity = driver.find_element_by_xpath(
                            "//table[@pl_prop_class='AMS-Data-Assets']/tbody/tr[2]/td[1]//span").text
                        TestResult.append("Record found after applying filter is : " + Deleted_Activity)
                        TestResultStatus.append("Pass")
                        time.sleep(2)

                        time.sleep(2)
                        if Deleted_Activity == "No items for the filters applied":
                            Search_Filter = driver.find_element_by_xpath(
                                "//th[@data-attribute-name='Asset']/div/span/a")
                            Search_Filter.click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            time.sleep(2)
                            driver.find_element_by_xpath("//a[@id='clearFilter']").click()
                            sheet2.cell(3, 3).value = None
                            wb2.save(loc2)
                        else:
                            # -----------Clicking on delete icon-----------
                            driver.find_element_by_xpath(
                                "//table[@pl_prop_class='AMS-Data-Activities']/tbody//tr[2]/td[5]//i[@data-test-id='202203300855020903290']").click()
                            # --------Clicking on submit button to confirm delete------------
                            driver.find_element_by_xpath("//button[@id='ModalButtonSubmit']").click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break

                        # ----------Verifying deleted asset record by applying filter again-----
                        try:
                            Search_Filter.click()
                            Search_Text_Field.send_keys(sheet2.cell(3, 1).value)
                            Filter_Apply_Btn.click()
                            for load in range(LONG_TIMEOUT):
                                try:
                                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                        time.sleep(0.5)
                                except Exception:
                                    break
                            Deleted_Activity2 = Deleted_Activity.text
                            if Deleted_Activity2 == "No items for the filters applied":
                                TestResult.append(
                                    "Asset type deleted successfully from asset type table records")
                                TestResultStatus.append("Pass")
                                TestResult.append(
                                    PageName + " is working fine. Asset type found after applying filter is: " + Deleted_Activity2)
                                TestResultStatus.append("Pass")
                                Search_Filter.click()
                                for load in range(LONG_TIMEOUT):
                                    try:
                                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                            time.sleep(0.5)
                                    except Exception:
                                        break
                                driver.find_element_by_xpath("//a[@id='clearFilter']").click()
                        except Exception:
                            pass
                        sheet2.cell(3, 3).value = None
                        wb2.save(loc2)
                except Exception:
                    pass
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify pagination clicks in Activity Table-----------------------------
            PageName = "Activity table Pagination"
            try:
                Page_Count3 = driver.find_element_by_xpath("//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']/tbody/tr/td[6]//label").text
                NumOfClicks3 = int(Page_Count3)
                print(Page_Count3)
                print(NumOfClicks3)

                counter2 = 0
                for i2 in range(1, NumOfClicks3 + 1):
                    AssNumRows3 = driver.find_elements_by_xpath(
                        "//table[@pl_prop_class='AMS-Data-Activities']/tbody/tr")
                    NumRows3 = len(AssNumRows3)

                    for ii2 in range(NumRows3):
                        if ii2 >= 5:
                            pass
                        else:
                            try:
                                Activity_Text = driver.find_element_by_xpath(
                                    "//table[@pl_prop_class='AMS-Data-Activities']/tbody/tr[" + str(
                                        ii2 + 2) + "]/td[1]//span").is_displayed()
                                time.sleep(0.5)
                                if Activity_Text == True:
                                    counter2 = counter2 + 1
                            except Exception:
                                break
                    print("No. of records on page no. " + str(i2) + " are : " + str(counter2))
                    driver.find_element_by_xpath("//div[@data-test-id='202206080844470310974']//table[@id='grid-desktop-paginator']//td[7]//button").click()
                TestResult.append(PageName + "  is working fine for " + str(NumOfClicks3) + " pages and No. of records found are : " + str(counter2))
                TestResultStatus.append("Pass")
                sheet2.cell(3, 6).value = counter2
                wb2.save(loc2)
            except Exception as errr:
                print(errr)
                TestResult.append(PageName + " is not working")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ----------------------------------------------------------------------------------------------

        except Exception as err:
            print("----------")
            print(err)
            TestResult.append("AMS Actions page is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


