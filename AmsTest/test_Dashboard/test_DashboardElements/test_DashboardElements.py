import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_DashboardElements"
  description = "This test scenario is to verify all the Elements present at Dashboard"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_AmsActionsElements"
  global Exe
  Exe="Yes"
  Directory = 'test_Dashboard/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        # LOADING_ELEMENT_XPATH = "//div[@class='main-loader LoaderImageLogo']"
        try:
            #---------------------------Verify Title of page----------------------------
            PageName="Title of page"
            Ptitle1="Dashboard"
            try:
                driver.find_element_by_xpath("//li[@data-test-id='dynamic-nav-menu_2']").click()
                PageTitle1 = driver.find_element_by_xpath("//h1[text()='Dashboard']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present ("+PageTitle1+")")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName +" is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            #---------------------------------------------------------------------------------

            # ---------------------------Verify Organization Chart Tab----------------------------
            PageName = "Organization Chart Tab"
            Ptitle1 = "Organization Chart"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202207110634440926775']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Total Assets Tab----------------------------
            PageName = "Total Assets Tab"
            Ptitle1 = "Total Assets"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202207110710030991819']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Breakdown Report Tab----------------------------
            PageName = "Breakdown Report Tab"
            Ptitle1 = "Breakdown Report"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202207110715540213464']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Broken Assets Tab----------------------------
            PageName = "Broken Assets Tab"
            Ptitle1 = "Broken Assets"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202207110624490670362']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Working Assets Tab----------------------------
            PageName = "Working Assets Tab"
            Ptitle1 = "Working Assets"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202207110542370647216']").text
                assert Ptitle1 in PageTitle1, PageName + " not able to open"
                TestResult.append(PageName + "  is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ------Scrolling Page----------------------------------------------------------
            for scrolldown in range (1,10):
                time.sleep(2)
                try:
                    driver.find_element_by_xpath("//table[@id='bodyTbl_right']//tr[1]/th[1]//div[@class='cellIn ']").click()
                    break
                except Exception:
                    #ActionChains(driver).key_down(Keys.).perform()
                    print("Inside Excep")
                    ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
                    print("Page Down")
                    pass
            #---------------------------------------------------------------------------------

            # ---------------------------Verify Recent Activities Table headers----------------------------
            PageName = "Recent Activities table headers"
            TableHeaders = ['Time','Description','Performed -By']
            LengthOfTH = len(TableHeaders)
            for i in range(LengthOfTH):
                try:
                    PageTitle1 = driver.find_element_by_xpath("//table[@id='bodyTbl_right']//tr[1]/th["+str[i]+"]//div[@class='cellIn ']").text
                    assert TableHeaders[i] in PageTitle1, PageName + " not able to open"
                    TestResult.append(PageName + "  is present "+TableHeaders[i])
                    TestResultStatus.append("Pass")
                except Exception:
                    TestResult.append("'"+TableHeaders[i] + "'"" is not present in " + PageName)
                    TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


