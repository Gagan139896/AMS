import datetime
import math
import re
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform

from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import os


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_DashboardWorking"
  description = "This test scenario is to verify working of Elements present at Dashboard page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_DashboardWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_Dashboard/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = ("D:/AMS/AmsTest/test_AmsActions/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active

        try:
            # ---------------------------Verify Dashboard icon click----------------------------
            PageName = "Dashboard icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//li[@data-test-id='dynamic-nav-menu_2']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                driver.find_element_by_xpath("//input[@id='pyGridActivePage']").click()
                PageTitle1 = driver.find_element_by_xpath("//h1[text()='Dashboard']").text
                if PageTitle1 == "Dashboard":
                    TestResult.append(PageName + "  is clickable")
                    TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # #---------------------------Verify Number of Total Assets--------------------------------
            # PageName = "Number of total assets"
            # Ptitle1 = ""
            # try:
            #     NumOfTotalAssets = driver.find_element_by_xpath("//div[@data-test-id='202207110710030992974']//button").text
            #     ExcelNumOfTotalAssets = sheet2.cell(3, 5).value
            #     print(NumOfTotalAssets)
            #     time.sleep(2)
            #     print(ExcelNumOfTotalAssets)
            #     if ExcelNumOfTotalAssets == int(NumOfTotalAssets) :
            #         TestResult.append(PageName + " are verified successfully")
            #         TestResultStatus.append("Pass")
            #     else:
            #         TestResult.append(
            #             PageName + " are not verified. Because number of total assets on 'AMS actions page' are : " + str(ExcelNumOfTotalAssets) + " and Number of total assets found on 'Dashboard page' are : " + str(
            #                 NumOfTotalAssets))
            #         TestResultStatus.append("Fail")
            # except Exception as err1:
            #     print(err1)
            # print()
            # time.sleep(TimeSpeed)
            # #---------------------------------------------------------------------------------
            #
            # # ---------------------------Verify Number of Working Assets--------------------------------
            # PageName = "Number of working assets"
            # Ptitle1 = ""
            # try:
            #     NumOfWorkingAssets = driver.find_element_by_xpath("//div[@data-test-id='202207110545480825508']//button").text
            #     ExcelNumOfWorkingAssets = sheet2.cell(3, 7).value
            #     print(NumOfWorkingAssets)
            #     time.sleep(2)
            #     print(ExcelNumOfWorkingAssets)
            #     if ExcelNumOfWorkingAssets == int(NumOfWorkingAssets):
            #         TestResult.append(PageName + " are verified successfully")
            #         TestResultStatus.append("Pass")
            #     else:
            #         TestResult.append(PageName + " are not verified. Because number of working assets on 'AMS actions page' are : "+str(ExcelNumOfWorkingAssets)+ " and Number of working assets found on 'Dashboard page' are : "+str(NumOfWorkingAssets))
            #         TestResultStatus.append("Fail")
            # except Exception as err2:
            #     print(err2)
            # print()
            # time.sleep(TimeSpeed)
            # # ---------------------------------------------------------------------------------
            #
            # # ---------------------------Verify Number of Broken Assets--------------------------------
            # PageName = "Number of broken assets"
            # Ptitle1 = ""
            # try:
            #     NumOfBrokenAssets = driver.find_element_by_xpath(
            #         "//div[@data-test-id='202207110624490670651']//button").text
            #     ExcelNumOfBrokenAssets = sheet2.cell(3, 8).value
            #     print(NumOfBrokenAssets)
            #     time.sleep(2)
            #     print(ExcelNumOfBrokenAssets)
            #     if ExcelNumOfBrokenAssets == int(NumOfBrokenAssets):
            #         TestResult.append(PageName + " are verified successfully")
            #         TestResultStatus.append("Pass")
            #     else:
            #         TestResult.append(PageName + " are not verified. Because number of broken assets on 'AMS actions page' are : " + str(
            #                 ExcelNumOfBrokenAssets) + " and number of broken assets found on Dashboard are : " + str(NumOfBrokenAssets))
            #         TestResultStatus.append("Fail")
            # except Exception as err3:
            #     print(err3)
            # print()
            # time.sleep(TimeSpeed)
            # # ---------------------------------------------------------------------------------
            #
            # # ------Scrolling Page----------------------------------------------------------
            # for scrolldown in range(1, 10):
            #     time.sleep(2)
            #     try:
            #         driver.find_element_by_xpath(
            #             "//table[@id='grid-desktop-paginator']//tr/td[7]//button")
            #         break
            #     except Exception:
            #         # ActionChains(driver).key_down(Keys.).perform()
            #         print("Inside Excep")
            #         ActionChains(driver).key_down(Keys.PAGE_DOWN).perform()
            #         print("Page Down")
            #         pass
            # # -----------------------------------------------------------------------------------------------
            #
            # # ---------------------------Verify pagination clicks in Recent activities Table-----------------------------
            # PageName = "Recent activities table pagination"
            #
            # try:
            #     # -------Getting total number of pages in Recent activities table-----
            #     Page_Count1 = driver.find_element_by_xpath(
            #         "//label[@data-test-id='20141007100658002115508']").text
            #     NumOfClicks1 = int(Page_Count1)
            #     print("Page_Count1 "+str(Page_Count1))
            #     print("NumOfClicks1 "+str(NumOfClicks1))
            #
            #     # ----For loop to page clicks----
            #     for i in range(1, NumOfClicks1):
            #         driver.find_element_by_xpath("//table[@id='grid-desktop-paginator']//tr/td[7]//button").click()
            #         time.sleep(1)
            #         ActivePageNum = driver.find_element_by_xpath("//input[@id='pyGridActivePage']").get_attribute('value')
            #         ActivePageNum = int(ActivePageNum)
            #         if ActivePageNum == NumOfClicks1:
            #             print("Recent activities table pagination is working")
            #             TestResult.append(PageName + " is working fine for " + str(NumOfClicks1)+ " pages")
            #             TestResultStatus.append("Pass")
            #         else:
            #             pass
            # except Exception as err:
            #     print(err)
            #     TestResult.append(PageName + " is not working")
            #     TestResultStatus.append("Fail")
            # print()
            # # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Organization chart tab click--------------------------------
            # PageName = "Organization chart tab"
            # Ptitle1 = ""
            # try:
            #     driver.find_element_by_xpath("//div[@data-test-id='202207110634440928719']//button").click()
            #     time.sleep(2)
            #     Org_Chart_Title = driver.find_element_by_xpath("//span[@id='modaldialog_hd_title']").text
            #     print(Org_Chart_Title)
            #     if Org_Chart_Title in "        Organization Chart       ":
            #         TestResult.append(PageName + " is clickable")
            #         TestResultStatus.append("Pass")
            # except Exception as err3:
            #     TestResult.append(PageName + " is not clickable")
            #     TestResultStatus.append("Fail")
            #     print(err3)
            # print()
            # time.sleep(TimeSpeed)
            # # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Operators List click--------------------------------
            # PageName = "Operators List"
            # Ptitle1 = ""
            # try:
            #     driver.find_element_by_xpath("//div[@data-layout-id='202209230519300692']").click()
            #     time.sleep(2)
            #     Full_Name = driver.find_element_by_xpath("//div[text()='Full Name']").is_displayed()
            #     print("Full_Name is "+str(Full_Name))
            #     if Full_Name == True:
            #         TestResult.append(PageName + " is clickable")
            #         TestResultStatus.append("Pass")
            # except Exception as err3:
            #     TestResult.append(PageName + " is not clickable")
            #     TestResultStatus.append("Fail")
            #     print(err3)
            # print()
            # time.sleep(TimeSpeed)
            # # ---------------------------------------------------------------------------------
            #
            # # ---------------------------Verify pagination clicks in Operators List Table-----------------------------
            # PageName = "Operators List table pagination"
            #
            # try:
            #     # -------Getting total number of pages in Operators List table-----
            #     Page_Count2 = driver.find_element_by_xpath("//div[@data-test-id='202207120559580902804']//label[@data-test-id='20141007100658002115508']").text
            #     NumOfClicks2 = int(Page_Count2)
            #     print("Page_Count2 " + str(Page_Count2))
            #     print("NumOfClicks2 " + str(NumOfClicks2))
            #
            #     # ----For loop to page clicks----
            #     for i in range(1, NumOfClicks2+1):
            #         Operators_List_Rows = driver.find_elements_by_xpath(
            #             "//table[@data-test-id='202207120559580902804-layout']/tbody//table//tr/td[1]/div")
            #         NumRows2 = int(len(Operators_List_Rows))
            #         print("NumRows1 " + str(NumRows2))
            #         for ii in range(NumRows2):
            #             Is_Present = driver.find_element_by_xpath(
            #                 "//table[@data-test-id='202207120559580902804-layout']/tbody//table//tr[" + str(
            #                     ii+2) + "]/td[1]/div").is_displayed()
            #             if Is_Present == True:
            #                 time.sleep(0.5)
            #                 User_Name = driver.find_element_by_xpath("//table[@data-test-id='202207120559580902804-layout']/tbody//table//tr[" + str(
            #                         ii+2) + "]/td[1]/div").text
            #                 print(User_Name)
            #                 #print(sheet2.cell(7, 1).value)
            #                 if User_Name == sheet2.cell(7, 1).value:
            #                     print("----------------------------------")
            #                     print("User_Name is " + str(User_Name)+ " and Excel value is "+str(sheet2.cell(7, 1).value))
            #                     print("Operators List table pagination is working")
            #                     print("----------------------------------")
            #                     TestResult.append(PageName + " is working fine for " + str(NumOfClicks2) + " pages and ")
            #                     TestResultStatus.append("Pass")
            #             else:
            #                 pass
            #         driver.find_element_by_xpath("//div[@data-test-id='202207120559580902804']//button[@title='Next Page']").click()
            #         time.sleep(0.5)
            #     driver.find_element_by_xpath("//div[@id='modaldialog_hd']/button[@class='container-close']").click()
            #     for load in range(LONG_TIMEOUT):
            #         try:
            #             if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
            #                 time.sleep(0.5)
            #         except Exception:
            #             break
            #     PageTitle1 = driver.find_element_by_xpath("//h1[text()='Dashboard']").text
            #     if PageTitle1 == "Dashboard":
            #         print("Close button is working fine on Organization chart page")
            #         TestResult.append("Close button is working fine on Organization chart page")
            #         TestResultStatus.append("Pass")
            # except Exception as err:
            #     print(err)
            #     TestResult.append(PageName + " is not working")
            #     TestResultStatus.append("Fail")
            # print()
            # # ---------------------------------------------------------------------------------

            # ---------------------------Verify Breakdown report tab click--------------------------------
            PageName = "Breakdown report tab"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//div[@data-test-id='202207110715540213271']//button").click()
                time.sleep(2)
                Breakdown_Case= driver.find_element_by_xpath("//h1[text()='Breakdown Cases']").is_displayed()
                print("Breakdown_Case is " + str(Breakdown_Case))
                if Breakdown_Case == True:
                    TestResult.append(PageName + " is clickable")
                    TestResultStatus.append("Pass")
            except Exception as err3:
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
                print(err3)
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Breakdown report table pagination and searching Breakdown id in table-----------------------------
            PageName = "Breakdown report table pagination"

            try:
                # -------Getting total number of pages in Breakdown report table-----
                Page_Count3= driver.find_element_by_xpath(
                    "//label[@data-test-id='20141007071054030615792']").text
                NumOfClicks3 = int(Page_Count3)
                print("Page_Count3 " + str(Page_Count3))
                print("NumOfClicks3 " + str(NumOfClicks3))

                counter = 0
                # ----For loop to page clicks----
                for i1 in range(1, NumOfClicks3):
                    Operators_List_Rows = driver.find_elements_by_xpath("//table[@data-test-id='201712290432230969267-layout']/tbody//table//tr/td[1]/div")
                    NumRows3 = int(len(Operators_List_Rows))
                    print("----------------------------------------------------------------")
                    print("Page No. " +str(i1)+ " and No. of records are "+str(NumRows3))
                    for ii1 in range(NumRows3):
                        time.sleep(0.25)
                        Is_Present = driver.find_element_by_xpath("//table[@data-test-id='201712290432230969267-layout']/tbody//table//tr[" + str(ii1 + 2) + "]/td[1]/div").is_displayed()
                        if Is_Present == True:
                            try:
                                Bd_Id_Text = driver.find_element_by_xpath(
                                    "//table[@data-test-id='201712290432230969267-layout']/tbody//table//tr[" + str(ii1 + 2) + "]/td[1]/div").text
                                counter = counter + 1
                                print("Records on page no. "+ str(i1)+ " " + str(Bd_Id_Text))
                                # print(sheet2.cell(7, 1).value)
                                if Bd_Id_Text == sheet2.cell(11, 1).value:
                                    print("----------------------------------")
                                    print("Bd_Id_Text is " + str(Bd_Id_Text) + " and Excel value is " + str(
                                        sheet2.cell(11, 1).value))
                                    print("Breakdown report table pagination is working")
                                    print("----------------------------------")
                                    TestResult.append(PageName + " is working fine for " + str(NumOfClicks3) + " pages and Number of records found in breakdown cases list are : " +str(counter))
                                    TestResultStatus.append("Pass")
                                    TestResult.append("'Breakdown id' generated during 'Breakdown Case' creation is verified in 'Breakdown cases list' on 'Dashboard page' ")
                                    TestResultStatus.append("Pass")
                            except Exception:
                                pass
                        else:
                            pass
                    time.sleep(0.25)
                    Page_Num = driver.find_element_by_xpath("//input[@id='pyGridActivePage']")
                    Page_Num.clear()
                    Page_Num.send_keys(i1+1)
                    time.sleep(0.25)
                    Page_Num.send_keys(Keys.ENTER)
                    #driver.find_element_by_xpath("//div[@data-test-id='201712290432230969267']//button[@class='nextPage']").click()
                driver.find_element_by_xpath("//button[@data-test-id='202207150314330815575']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                PageTitle1 = driver.find_element_by_xpath("//h1[text()='Dashboard']").text
                if PageTitle1 == "Dashboard":
                    print("Close button is working fine on Organization chart page")
                    TestResult.append("Close button is working fine on Breakdown report page")
                    TestResultStatus.append("Pass")
            except Exception as err:
                print(err)
                TestResult.append(PageName + " is not working")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

        except Exception as errr:
            TestResult.append("Dashboard page is not working fine. Below error is found "+str(errr))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


