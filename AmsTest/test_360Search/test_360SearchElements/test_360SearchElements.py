import datetime
import math
import os
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_360SearchElements"
  description = "This test scenario is to verify all the Elements present at 360 Search"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_360SearchElements"
  global Exe
  Exe="Yes"
  Directory = 'test_360Search/'

  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  print(MachineName)
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")
      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    global Element1
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 2
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        try:
            # -------360 Search icon click------------------------
            PageName = "360 Search icon"
            Ptitle1 = "GO FOR 360° SEARCH"
            try:
                driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[1]/a").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202205040525180362469']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present in left menu and able to click")
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                print(PageName + " is not clickable")
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            time.sleep(3)
            # --------------------------------------------------------------------------------------------

            # ---------------------------Verify Page title-----------------------------
            PageName = "Page title"
            Ptitle1 = "GO FOR 360° SEARCH"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@data-test-id='202205040525180362469']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " ("+PageTitle1+") is present on 360 search page")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present on 360 search page")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify presence of Search text-------------------------------------
            PageName = "Search text"
            Ptitle1 = "Search"
            try:
                PageTitle1 = driver.find_element_by_xpath("//h2[text()='Search']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Asset Type field label -------------------------------------
            PageName = "Asset Type field label"
            Ptitle1 = "Asset Type"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//label[@data-test-id='202204250843470340998-Label']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Asset field label -------------------------------------
            PageName = "Asset field label"
            Ptitle1 = "Asset"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//label[@data-test-id='202204250843470341228-Label']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Activity field label -------------------------------------
            PageName = "Activity field label"
            Ptitle1 = "Activity"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//label[@data-test-id='202204250851440995455-Label']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Search button -------------------------------------
            PageName = "Search button"
            Ptitle1 = "Search"
            try:
                PageTitle1 = driver.find_element_by_xpath("//button[@data-test-id='202206091157570013179']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Clear button -------------------------------------
            PageName = "Clear button"
            Ptitle1 = "Clear"
            try:
                PageTitle1 = driver.find_element_by_xpath("//button[@data-test-id='202206091157570014628']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Activities tab-------------------------------------
            PageName = "Activities tab"
            Ptitle1 = "Activities"
            try:
                PageTitle1 = driver.find_element_by_xpath(
                    "//div[@data-layout-id='202209150520030684']/h3/i").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Asset type-------------------------------------
            PageName = "Asset type"
            Ptitle1 = "Asset type"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-layout-id='202209150520030801']/h3/i").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Presence of Assets-------------------------------------
            PageName = "Assets"
            Ptitle1 = "Assets"
            try:
                PageTitle1 = driver.find_element_by_xpath("//div[@data-layout-id='202209150520030978']/h3/i").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " not present"
                TestResult.append(PageName + " is present")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Presence of elements in 360 search-Activities table-----------------------------
            inside = "360 search-Activities"
            # ---------------loop for Columns in table for table headers View----------
            ItemList = ["Name", "Asset Type", "Frequency", "Planning In Days", "Region", "State", "Unit"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                print(ii)
                try:
                    Element1 = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='1']/div[2]//tbody/tr[1]/th["+str(ii+2)+"]/div[1]").text
                    print(Element1)
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under [ " + inside + " ] table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under [ " + inside + " ] table\n" + ListD)
                TestResultStatus.append("Fail")
            driver.find_element_by_xpath("//h3[text()='Asset Type']").click()
            for load in range(LONG_TIMEOUT):
                try:
                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                        time.sleep(0.5)
                except Exception:
                    break
            # # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Presence of elements in 360 search-Asset type table-----------------------------
            inside = "360 search-Asset type"
            # ---------------loop for Columns in table for table headers View----------
            ItemList = ["Asset Type", "Region",  "State", "Unit"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                print(ii)
                try:
                    Element1 = driver.find_element_by_xpath(
                        "//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='2']/div[2]//tbody/tr[1]/th[" + str(ii + 2) + "]/div[1]").text
                    print(Element1)
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under [ " + inside + " ] table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under [ " + inside + " ] table\n" + ListD)
                TestResultStatus.append("Fail")
            driver.find_element_by_xpath("//h3[text()='Assets']").click()
            for load in range(LONG_TIMEOUT):
                try:
                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                        time.sleep(0.5)
                except Exception:
                    break
            # # ---------------------------------------------------------------------------------

            # # ---------------------------Verify Presence of elements in 360 search-Asset table-----------------------------
            inside = "360 search-Asset"
            # ---------------loop for Columns in table for table headers View----------
            ItemList = ["Asset Name", "Linked to Asset Type", "Whether In Use", "Specifications", "Service Start Date", "Region",  "State", "Unit"]
            ItemPresent = []
            ItemNotPresent = []
            for ii in range(len(ItemList)):
                Text1 = ItemList[ii]
                print(ii)
                try:
                    Element1 = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='3']/div[2]//tbody/tr[1]/th[" + str(
                            ii + 2) + "]/div[1]").text
                    print(Element1)
                except Exception:
                    pass
                try:
                    assert Text1 in Element1, Text1 + " column under " + inside + " table is not present"
                    ItemPresent.append(Text1)
                except Exception as e1:
                    ItemNotPresent.append(Text1)
            if ItemPresent:
                print("ItemPresent list is not empty")
                ListC = ', '.join(ItemPresent)
                TestResult.append("Below columns are present under [ " + inside + " ] table\n" + ListC)
                TestResultStatus.append("Pass")
            if ItemNotPresent:
                print("ItemNotPresent list is not empty")
                ListD = ', '.join(ItemNotPresent)
                TestResult.append("Below columns are not present under [ " + inside + " ] table\n" + ListD)
                TestResultStatus.append("Fail")

            # # ---------------------------------------------------------------------------------
        except Exception as err:
            print(err)
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


