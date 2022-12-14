import datetime
import math
import re
from selenium.webdriver.support.select import Select
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import os
import random
import string


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(password)


@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_360SearchWorking"
  description = "This test scenario is to verify the Working of 360 Search"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_360SearchWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_360Search/'

  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  print(MachineName)
  if MachineName=="DESKTOP-JLLTS65":
      path=path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="D:/AMS/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      #driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    global select
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//div[@class='loader']"
        loc2 = ("D:/AMS/AmsTest/test_AmsActions/test_AmsActionsWorking/DataRecord.xlsx")
        wb2 = openpyxl.load_workbook(loc2)
        sheet2 = wb2.active
        try:
            #-------360 Search icon click------------------------
            PageName = "360 Search icon"
            Ptitle1 = "GO FOR 360?? SEARCH"
            try:
                driver.find_element_by_xpath("//div[@data-test-id='201808081157350664772']/div[2]//li[1]/a").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                driver.find_element_by_xpath("//div[@data-layout-id='202209150520030801']").click()
                PageTitle1 = driver.find_element_by_xpath("//div[@data-test-id='202205040525180362469']").text
                assert PageTitle1 in Ptitle1, PageName + " not present"
                print(PageName + " is present in left menu and able to click")
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception:
                print(PageName + " is not clickable")
                TestResult.append(PageName + " is not clickable")
                TestResultStatus.append("Fail")
            time.sleep(3)
            #--------------------------------------------------------------------------------------------

            #----------------Searching asset type---------------------------------------------
            PageName = "360 search for Asset type"
            try:
                #--------Entering asset type------------
                AssetType = driver.find_element_by_xpath("//input[@data-test-id='202204250843470340998']")
                AssetType.send_keys(sheet2.cell(3, 1).value)
                time.sleep(2)
                AssetType.send_keys(Keys.DOWN)
                AssetType.send_keys(Keys.ENTER)
                time.sleep(3)
                #---------Clicking on Search button------------
                driver.find_element_by_xpath("//button[@data-test-id='202206091157570013179']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                #-----Clicking on Asset type tab-----------
                driver.find_element_by_xpath("//h3[text()='Asset Type']").click()
                time.sleep(2)
                try:
                    #-------Getting text of Filtered Asset type---------
                    Asset_Type_Text = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='2']/div[2]//tbody/tr[2]/td[2]/span").text
                    if sheet2.cell(3, 1).value == Asset_Type_Text:
                        print(PageName + " is working fine")
                        TestResult.append(PageName + " is working fine")
                        TestResultStatus.append("Pass")
                except Exception:
                    time.sleep(2)
                    #--------Getting text when no asset type record found after applying filter-----------
                    No_Results = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='2']/div[2]//tbody/tr[2]/td[1]").is_displayed()
                    if No_Results == True:
                        print(PageName + " is not working fine. Because asset type not found in grid after applying search")
                        TestResult.append(PageName + " is not working fine. Because asset type not found in grid after applying search")
                        TestResultStatus.append("Fail")
                driver.find_element_by_xpath("//button[@data-test-id='202206091157570014628']").click()
                try:
                   Is_True = AssetType.get_attribute('value')
                   if Is_True == sheet2.cell(3, 1).value:
                       print("Clear button is not working fine for Asset type search")
                       TestResult.append("Clear button is not working fine for Asset type search")
                       TestResultStatus.append("Fail")
                except Exception:
                    print("Clear button is working fine for Asset type search")
                    TestResult.append("Clear button is working fine for Asset type search")
                    TestResultStatus.append("Pass")
            except Exception as err:
                print("Asset type search not is working fine. Below error found : "+str(err))
                TestResult.append("Asset type search not is working fine. Below error found : "+str(err))
                TestResultStatus.append("Fail")
            time.sleep(3)
            #---------------------------------------------------------------------------------------

            # ----------------Searching Asset---------------------------------------------
            PageName = "360 search for Asset"
            try:
                # --------Entering Asset------------
                Asset = driver.find_element_by_xpath("//input[@data-test-id='202204250843470341228']")
                Asset.send_keys(sheet2.cell(3, 2).value)
                time.sleep(2)
                Asset.send_keys(Keys.DOWN)
                Asset.send_keys(Keys.ENTER)
                time.sleep(3)
                # ---------Clicking on Search button------------
                driver.find_element_by_xpath("//button[@data-test-id='202206091157570013179']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                # -----Clicking on Asset tab-----------
                driver.find_element_by_xpath("//h3[text()='Assets']").click()
                time.sleep(2)
                try:
                    #----------Getting text of filtered asset---------------
                    Asset_Text = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='3']/div[2]//tbody/tr[2]/td[2]/span").text
                    if sheet2.cell(3, 2).value == Asset_Text:
                        print(PageName + " is working fine for asset type")
                        TestResult.append(PageName + " is working fine for asset type")
                        TestResultStatus.append("Pass")
                except Exception:
                    time.sleep(2)
                    #-------------Getting text when No Asset Record found after applying filter-----------
                    No_Results = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='3']/div[2]//tbody/tr[2]/td[1]").is_displayed()
                    if No_Results == True:
                        print(PageName + " is not working fine for asset. Because asset not found in grid after applying search")
                        TestResult.append(PageName + " is not working fine for asset. Because asset not found in grid after applying search")
                        TestResultStatus.append("Fail")
                driver.find_element_by_xpath("//button[@data-test-id='202206091157570014628']").click()
                try:
                   Is_True = Asset.get_attribute('value')
                   if Is_True == sheet2.cell(3, 2).value:
                       print("Clear button is not working fine for Asset search")
                       TestResult.append("Clear button is not working fine for Asset search")
                       TestResultStatus.append("Fail")
                except Exception:
                    print("Clear button is working fine for Asset search")
                    TestResult.append("Clear button is working fine for Asset search")
                    TestResultStatus.append("Pass")
            except Exception as err:
                print("Asset search is not working fine. Below error found : " + str(err))
                TestResult.append("Asset search not is working fine. Below error found : " + str(err))
                TestResultStatus.append("Fail")
            time.sleep(3)
            # ---------------------------------------------------------------------------------------

            # ----------------Searching Activity---------------------------------------------
            PageName = "360 search for Activity"
            try:
                # --------Entering Activity------------
                Activity = driver.find_element_by_xpath("//input[@data-test-id='202204250851440995455']")
                Activity.send_keys(sheet2.cell(3, 3).value)
                time.sleep(2)
                Activity.send_keys(Keys.DOWN)
                Activity.send_keys(Keys.ENTER)
                time.sleep(3)
                # ---------Clicking on Search button------------
                driver.find_element_by_xpath("//button[@data-test-id='202206091157570013179']").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                try:
                    # ----------Getting text of filtered activity---------------
                    Activity_Text = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='1']/div[2]//tbody/tr[2]/td[2]/span").text
                    if sheet2.cell(3, 3).value == Activity_Text:
                        print(PageName + " is working fine for activity")
                        TestResult.append(PageName + " is working fine for activity")
                        TestResultStatus.append("Pass")
                except Exception:
                    time.sleep(2)
                    # -------------Getting text when No activity Record found after applying filter-----------
                    No_Results = driver.find_element_by_xpath("//div[@data-test-id='202206090831230119525']//div[@data-lg-child-id='1']/div[2]//tbody/tr[2]/td[1]").is_displayed()
                    if No_Results == True:
                        print(PageName + " is not working fine for activity. Because activity not found in grid after applying search")
                        TestResult.append(PageName + " is not working fine for activity. Because activity not found in grid after applying search")
                        TestResultStatus.append("Fail")
                driver.find_element_by_xpath("//button[@data-test-id='202206091157570014628']").click()
                try:
                    Is_True = Activity.get_attribute('value')
                    if Is_True == sheet2.cell(3, 3).value:
                        print("Clear button is not working fine for Activity search")
                        TestResult.append("Clear button is not working fine for Activity search")
                        TestResultStatus.append("Fail")
                except Exception:
                    print("Clear button is working fine for Activity search")
                    TestResult.append("Clear button is working fine for Activity search")
                    TestResultStatus.append("Pass")
            except Exception as err:
                print("Activity search is not working fine. Below error found : " + str(err))
                TestResult.append("Activity search not is working fine. Below error found : " + str(err))
                TestResultStatus.append("Fail")
            time.sleep(3)
            # ---------------------------------------------------------------------------------------

        except Exception as rrr:
            print("360 search  is not working fine. Below error found : " + str(rrr))
            TestResult.append("360 search not is working fine. Below error found : " + str(rrr))
            TestResultStatus.append("Fail")
    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


