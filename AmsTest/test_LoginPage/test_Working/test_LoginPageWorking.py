import datetime
import os
import time
import openpyxl
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from sys import platform


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path
  global p

  TestName = "test_LoginPageWorking"
  description = "This test scenario is to verify working of Login Process"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_Working"
  global Exe
  Exe="Yes"

  Directory = 'test_LoginPage/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")

  yield
  if Exe == "Yes":
      time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now() + time_change
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/Logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 5
        LONG_TIMEOUT = 400
        LOADING_ELEMENT_XPATH = "//div[@id='appian-working-indicator-hidden']"
        UName="CountryHead_4"
        PName="Rules@12345"
        try:
            # ---------------------------Verify Validation messages for blank username field----------------------------
            PageName = "Validation message"
            Ptitle1 = "The information you entered was not recognized "
            try:
                driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(PName)
                driver.find_element_by_xpath("//button[@id='sub']/span").click()
                PageTitle1 = driver.find_element_by_xpath("//div[@id='error']").text
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + " is present when user click submit by entering Valid password only")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present when user click submit by entering Valid password only")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Validation messages for blank username and password----------------------------
            PageName = "Validation message"
            Ptitle1 = "The information you entered was not recognized "
            try:
                driver.find_element_by_xpath("//button[@id='sub']/span").click()
                PageTitle1 = driver.find_element_by_xpath("//div[@id='error']").text
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(PageName + " is present when user click submit without entering username and password")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not present when user click submit without entering username and password")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Validation messages for blank password field----------------------------
            PageName = "Validation message"
            Ptitle1 = "The information you entered was not recognized "
            try:
                driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(UName)
                driver.find_element_by_xpath("//button[@id='sub']/span").click()
                PageTitle1 = driver.find_element_by_xpath("//div[@id='error']").text
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(
                    PageName + " is present when user click submit by entering Valid Username only")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(
                    PageName + " is not present when user click submit by entering Valid Username only")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Validation messages for invalid credentials----------------------------
            PageName = "Validation message"
            Ptitle1 = "The information you entered was not recognized "
            try:
                driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys("abc")
                driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys("123")
                driver.find_element_by_xpath("//button[@id='sub']/span").click()
                PageTitle1 = driver.find_element_by_xpath("//div[@id='error']").text
                assert PageTitle1 in Ptitle1, PageName + " not able to open"
                TestResult.append(
                    PageName + " is present when user click submit by entering invalid credentials")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(
                    PageName + " is not present when user click submit by entering invalid credentials")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Username Field-----------------------------
            PageName = "Username"
            try:
                driver.find_element_by_xpath("//input[@id='txtUserID']").send_keys(UName)
                TestResult.append(PageName + " entered successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not able to enter")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Password Field-----------------------------
            PageName = "Password"
            try:
                driver.find_element_by_xpath("//input[@id='txtPassword']").send_keys(PName)
                TestResult.append(PageName + " entered successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not able to enter")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify SignIn Button-----------------------------
            PageName = "Sign In Button"
            try:
                driver.find_element_by_xpath("//button[@id='sub']/span").click()
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed()==True:
                            time.sleep(0.5)
                    except Exception:
                        break
                TestResult.append(PageName + " clicked successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not click")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # ---------------------------Verify Lost Password Link-----------------------------
            PageName = "Login process"
            Ptitle1 = "Asset Management System"
            try:
                PageTitle1 = driver.find_element_by_xpath("//h1[text()='Asset Management System']").text
                print(PageTitle1)
                assert PageTitle1 in Ptitle1, PageName + " is not perform"
                TestResult.append(PageName + " performed successfully")
                TestResultStatus.append("Pass")
            except Exception:
                TestResult.append(PageName + " is not perform")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------


        except Exception:
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


