import datetime
import math
import re
import time
import openpyxl
from datetime import datetime,date
import datetime as datetime
from fpdf import FPDF
import pytest
from selenium import webdriver
import allure
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from sys import platform
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pyperclip
from pathlib import Path
import os
import ntpath
import os.path

from setuptools import glob


@allure.step("Entering username ")
def enter_username(username):
  driver.find_element_by_id("email").send_keys(username)

@allure.step("Entering password ")
def enter_password(password):
  driver.find_element_by_id("password").send_keys(password)

@pytest.fixture()
def test_setup():
  global driver
  global TestName
  global description
  global TestResult
  global TestResultStatus
  global TestDirectoryName
  global path

  TestName = "test_MyTeam"
  description = "This test scenario is to verify the working og My Team page"
  TestResult = []
  TestResultStatus = []
  TestFailStatus = []
  FailStatus="Pass"
  TestDirectoryName = "test_MyTeamWorking"
  global Exe
  Exe="Yes"
  Directory = 'test_MyTeam/'
  if platform == "linux" or platform == "linux2":
      path = '/home/legion/office 1wayit/AVER/AverTest1/' + Directory
  elif platform == "win32" or platform == "win64":
      path = 'D:/AMS/AmsTest/' + Directory

  MachineName = os.getenv('COMPUTERNAME')
  if MachineName == "DESKTOP-JLLTS65":
      path = path.replace('D:', 'C:')

  ExcelFileName = "Execution"
  locx = (path+'Executiondir/' + ExcelFileName + '.xlsx')
  wbx = openpyxl.load_workbook(locx)
  sheetx = wbx.active

  for ix in range(1, 100):
      if sheetx.cell(ix, 1).value == None:
          break
      else:
          if sheetx.cell(ix, 1).value == TestName:
              if sheetx.cell(ix, 2).value == "No":
                  Exe="No"
              elif sheetx.cell(ix, 2).value == "Yes":
                  Exe="Yes"

  if Exe == "Yes":
      if platform == "linux" or platform == "linux2":
          driver = webdriver.Chrome(
              executable_path="/home/legion/office 1wayit/AVER/AverTest1/chrome/chromedriverLinux1")
      elif platform == "win32" or platform == "win64":
          driver = webdriver.Chrome(executable_path="/AmsTest/chrome/chromedriver.exe")

      driver.implicitly_wait(10)
      driver.maximize_window()
      driver.get("https://pegaqa.crochetech.com/prweb")
      enter_username("CountryHead_4")
      enter_password("Rules@12345")
      driver.find_element_by_xpath("//button[@id='sub']/span").click()

  yield
  if Exe == "Yes":
      #time_change = datetime.timedelta(hours=5)
      new_time = datetime.datetime.now()
      ctReportHeader = new_time.strftime("%d %B %Y %I %M%p")

      ct = new_time.strftime("%d_%B_%Y_%I_%M%p")

      class PDF(FPDF):
          def header(self):
              self.image(path+'EmailReportContent/logo.png', 10, 8, 33)
              self.set_font('Arial', 'B', 15)
              self.cell(73)
              self.set_text_color(0, 0, 0)
              self.cell(35, 10, ' Test Report ', 1, 1, 'B')
              self.set_font('Arial', 'I', 10)
              self.cell(150)
              self.cell(30, 10, ctReportHeader, 0, 0, 'C')
              self.ln(20)

          def footer(self):
              self.set_y(-15)
              self.set_font('Arial', 'I', 8)
              self.set_text_color(0, 0, 0)
              self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

      pdf = PDF()
      pdf.alias_nb_pages()
      pdf.add_page()
      pdf.set_font('Times', '', 12)
      pdf.cell(0, 10, "Test Case Name:  "+TestName, 0, 1)
      pdf.multi_cell(0, 10, "Description:  "+description, 0, 1)

      for i1 in range(len(TestResult)):
         pdf.set_fill_color(255, 255, 255)
         pdf.set_text_color(0, 0, 0)
         if (TestResultStatus[i1] == "Fail"):
             #print("Fill Red color")
             pdf.set_text_color(255, 0, 0)
             TestFailStatus.append("Fail")
         TestName1 = TestResult[i1].encode('latin-1', 'ignore').decode('latin-1')
         pdf.multi_cell(0, 7,str(i1+1)+")  "+TestName1, 0, 1,fill=True)
         TestFailStatus.append("Pass")
      pdf.output(TestName+"_" + ct + ".pdf", 'F')

      #-----------To check if any failed Test case present-------------------
      for io in range(len(TestResult)):
          if TestFailStatus[io]=="Fail":
              FailStatus="Fail"
      # ---------------------------------------------------------------------

      # -----------To add test case details in PDF details sheet-------------
      ExcelFileName = "FileName"
      loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
      wb = openpyxl.load_workbook(loc)
      sheet = wb.active
      print()
      check = TestName
      PdfName = TestName + "_" + ct + ".pdf"
      checkcount = 0

      for i in range(1, 100):
          if sheet.cell(i, 1).value == None:
              if checkcount == 0:
                  sheet.cell(row=i, column=1).value = check
                  sheet.cell(row=i, column=2).value = PdfName
                  sheet.cell(row=i, column=3).value = TestDirectoryName
                  sheet.cell(row=i, column=4).value = description
                  sheet.cell(row=i, column=5).value = FailStatus
                  checkcount = 1
              wb.save(loc)
              break
          else:
              if sheet.cell(i, 1).value == check:
                  if checkcount == 0:
                    sheet.cell(row=i, column=2).value = PdfName
                    sheet.cell(row=i, column=3).value = TestDirectoryName
                    sheet.cell(row=i, column=4).value = description
                    sheet.cell(row=i, column=5).value = FailStatus
                    checkcount = 1
      #----------------------------------------------------------------------------

      #---------------------To add Test name in Execution sheet--------------------
      ExcelFileName1 = "Execution"
      loc1 = (path+'Executiondir/' + ExcelFileName1 + '.xlsx')
      wb1 = openpyxl.load_workbook(loc1)
      sheet1 = wb1.active
      checkcount1 = 0

      for ii1 in range(1, 100):
          if sheet1.cell(ii1, 1).value == None:
              if checkcount1 == 0:
                  sheet1.cell(row=ii1, column=1).value = check
                  checkcount1 = 1
              wb1.save(loc1)
              break
          else:
              if sheet1.cell(ii1, 1).value == check:
                  if checkcount1 == 0:
                    sheet1.cell(row=ii1, column=1).value = check
                    checkcount1 = 1
      #-----------------------------------------------------------------------------

      driver.quit()

@pytest.mark.smoke
def test_VerifyAllClickables(test_setup):
    if Exe == "Yes":
        TimeSpeed = 2
        SHORT_TIMEOUT = 3
        LONG_TIMEOUT = 60
        LOADING_ELEMENT_XPATH = "//body[@class='sidebar-xs loader_overlay']"
        try:
            # ---------------------------Verify Downloads icon click-----------------------------
            PageName = "Downloads icon"
            Ptitle1 = ""
            try:
                driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
                time.sleep(2)

                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[13]/a").click()
                time.sleep(2)

                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                time.sleep(2)
                TestResult.append(PageName + " is present in left menu and able to click")
                TestResultStatus.append("Pass")
            except Exception as ee:
                print(ee)
                TestResult.append(PageName + " is not present")
                TestResultStatus.append("Fail")
            print()
            time.sleep(TimeSpeed)
            # ---------------------------------------------------------------------------------

            # # ---------------------------Verify working of Back button on Downloads page -----------------------------
            # PageName = "Back button"
            # Ptitle1 = "Rae"
            # try:
            #     driver.find_element_by_xpath("//a[text()='Back']").click()
            #     for load in range(LONG_TIMEOUT):
            #         try:
            #             if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
            #                 time.sleep(0.5)
            #         except Exception:
            #             break
            #     time.sleep(2)
            #     PageTitle1 = driver.find_element_by_xpath("//div[@class='hed_wth_srch']/h2").text
            #     print(PageTitle1)
            #     assert PageTitle1 in Ptitle1, PageName + " not present"
            #     TestResult.append(PageName + " is clickable")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " is not clickable")
            #     TestResultStatus.append("Fail")
            # print()
            # time.sleep(TimeSpeed)
            # # ---------------------------------------------------------------------------------
            #
            # # ----------------Verify Downloads icon click after verifying back--------
            # PageName = "Downloads icon"
            # Ptitle1 = ""
            # try:
            #     driver.find_element_by_xpath("//i[@class='icon-paragraph-justify3']/parent::a").click()
            #     time.sleep(2)
            #     driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[13]/a").click()
            #     time.sleep(2)
            #
            #     for load in range(LONG_TIMEOUT):
            #         try:
            #             if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
            #                 time.sleep(0.5)
            #         except Exception:
            #             break
            #     TestResult.append(PageName + "  is opened again after verifying back button")
            #     TestResultStatus.append("Pass")
            # except Exception:
            #     TestResult.append(PageName + " is not opened again after verifying back button")
            #     TestResultStatus.append("Fail")
            # print()
            # time.sleep(TimeSpeed)
            # for load in range(LONG_TIMEOUT):
            #     try:
            #         if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
            #             time.sleep(0.5)
            #     except Exception:
            #         break
            # # ---------------------------------------------------------------------------------

            # --------------Generating ABA file from invoice entry table-------------------------------------
            try:
                print()
                #------------------Clicking on invoice entry icon-----------------------------
                driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[4]/a").click()
                time.sleep(1)
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                TestResult.append("Clicking on Invoice entry icon")
                TestResultStatus.append("Pass")

                #------------------Selecting reference number checkbox-----------------------------
                driver.find_element_by_xpath("//tbody[@id='invoiceEntryListingAjaxView']/tr[1]/td[1]/input").click()
                time.sleep(1)
                TestResult.append("Invoice reference number is selected from invoice entry table")
                TestResultStatus.append("Pass")
                # ------------------Clicking on action button-----------------------------
                driver.find_element_by_xpath("//button[@id='actionBtn']").click()
                time.sleep(1)
                TestResult.append("Clicking on action button")
                TestResultStatus.append("Pass")
                driver.find_element_by_xpath("//button[@id='actionBtn']").click()
                time.sleep(1)
                # ------------------Selecting generate ABA file option-----------------------------
                driver.find_element_by_xpath("//a[text()='Generate ABA File']").click()
                TestResult.append("Selecting generate ABA file option under action button")
                TestResultStatus.append("Pass")
                for load in range(LONG_TIMEOUT):
                    try:
                        if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                            time.sleep(0.5)
                    except Exception:
                        break
                TestResult.append("ABA file is generated successfully")
                TestResultStatus.append("Pass")

            except Exception as it:
                print("Generate ABA file process not completed due to below error : "+str(it))
                TestResult.append("Generate ABA file process not completed due to below error : "+str(it))
                TestResultStatus.append("Fail")
                pass

            #-------------------------------------------------------------------------------------


            # --------------Finding latest downloaded file in downloads folder-------------------------------------
            TestResult.append("Searching downloaded file in downloads folder")
            TestResultStatus.append("Pass")
            time.sleep(3)
            folder_path = str(Path.home() / "Downloads")
            file_type = r'\*'
            files = glob.glob(folder_path + file_type)
            max_file = max(files, key=os.path.getctime)
            print(max_file)
            filename = ntpath.basename("'r'" + str(max_file))
            print(filename)
            TestResult.append("Downloaded ABA file is found in downloads folder. The file name is : \n"+str(filename))
            TestResultStatus.append("Pass")

            driver.find_element_by_xpath("//div[@class='card card-sidebar-mobile']/ul/li[13]/a").click()
            TestResult.append("Downloads icon is clicked again after generating ABA file to verify the downloaded file in ")
            TestResultStatus.append("Pass")
            for load in range(LONG_TIMEOUT):
                try:
                    if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                        time.sleep(0.5)
                except Exception:
                    break

            # -------------------------------------------------------------------------------------

            # --------------Verifying pagination clicks for downloads listing table-------------------------------------
            RecordsPerPage = 50
            TotalItem = driver.find_element_by_xpath("//div[@id='table_data_info']").text
            print(TotalItem)

            substr = "of"
            x = TotalItem.split(substr)
            string_name = x[0]
            TotalItemAfterOf = x[1]
            abc = ""
            countspace = 0
            for element in range(0, len(string_name)):
                if string_name[(len(string_name) - 1) - element] == " ":
                    countspace = countspace + 1
                    if countspace == 2:
                        break
                else:
                    abc = abc + string_name[(len(string_name) - 1) - element]
            abc = abc[::-1]
            TotalItemBeforeOf = abc
            TotalItemAfterOf = TotalItemAfterOf.split(" ")
            TotalItemAfterOf = TotalItemAfterOf[1]
            TotalItemAfterOf = re.sub('[^A-Za-z0-9]+', '', TotalItemAfterOf)

            TotalItemAfterOf = int(TotalItemAfterOf)
            TotalPages = TotalItemAfterOf / RecordsPerPage
            NumberOfPages = math.ceil(float(TotalPages))
            print(TotalItemAfterOf)
            print(NumberOfPages)
            print("RecordsPerPage is " + str(RecordsPerPage))

            for i in range(NumberOfPages):
                if i == NumberOfPages - 1:
                    TestResult.append(
                        "Pagination for [ " + str(TotalItemAfterOf) + " ] no. of records is successfully verified")
                    TestResultStatus.append("Pass")
                    break
                ItemLength = driver.find_elements_by_xpath("//table[@id='table_data']/tbody/tr")
                ItemLength = len(ItemLength)
                print(ItemLength)
                for ii in range(ItemLength):
                    Text1 = driver.find_element_by_xpath("//table[@id='table_data']/tbody/tr["+str(ii+1)+"]/td[6]").text
                    if Text1 == filename:
                        print("Downloaded file is found in downloads section of application and verified successfully")
                        TestResult.append(
                            "Downloaded file is found in downloads section of application and verified successfully")
                        TestResultStatus.append("Pass")
                        driver.find_element_by_xpath("//table[@id='table_data']/tbody/tr[" + str(ii + 1) + "]/td[9]/a").click()
                        time.sleep(2)
                        driver.find_element_by_xpath("//button[text()='Yes']").click()
                        for load in range(LONG_TIMEOUT):
                            try:
                                if driver.find_element_by_xpath(LOADING_ELEMENT_XPATH).is_displayed() == True:
                                    time.sleep(0.5)
                            except Exception:
                                break
                        TestResult.append(
                            "File removed from the downloads section of application")
                        TestResultStatus.append("Pass")

                        #------------Removing downloaded file from the system--------------------------
                        os.remove(max_file)
                        TestResult.append(
                            "Downloaded file removed from the system")
                        TestResultStatus.append("Pass")
                        break
                    else:
                        print("Downloaded file is not found in downloads section of application")
                        TestResult.append(
                            "Downloaded file is not found in downloads section of application")
                        TestResultStatus.append("Fail")
                    time.sleep(0.5)

                driver.find_element_by_xpath("//div[@class='dataTables_paginate paging_simple_numbers']/a[2]").click()
                time.sleep(2)
            if i != NumberOfPages - 1:
                TestResult.append(
                    "Pagination for [ " + str(TotalItemAfterOf) + " ] no. of records is not working correctly")
                TestResultStatus.append("Fail")

            # ---------------------------------------------------------------------------------

        except Exception as err:
            print(err)
            TestResult.append("ABA file verification process is not working correctly. Below error found\n"+str(err))
            TestResultStatus.append("Fail")
            pass

    else:
        print()
        print("Test Case skipped as per the Execution sheet")
        skip = "Yes"

        # -----------To add Skipped test case details in PDF details sheet-------------
        ExcelFileName = "FileName"
        loc = (path+'PDFFileNameData/' + ExcelFileName + '.xlsx')
        wb = openpyxl.load_workbook(loc)
        sheet = wb.active
        check = TestName

        for i in range(1, 100):
            if sheet.cell(i, 1).value == check:
                sheet.cell(row=i, column=5).value = "Skipped"
                wb.save(loc)
        # ----------------------------------------------------------------------------


